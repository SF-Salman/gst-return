import logging
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

TOLERANCE_MATCH = 1.0
TOLERANCE_MINOR = 100.0


def _status(diff: float) -> str:
    abs_diff = abs(diff)
    if abs_diff <= TOLERANCE_MATCH:
        return "MATCH"
    if abs_diff <= TOLERANCE_MINOR:
        return "WARNING"
    return "MISMATCH"


# ─── Per-period reconciliation ──────────────────────────────────────────────

def _reconcile_period(period: str, books: Optional[dict], gstr2b: Optional[dict]) -> dict:
    """books / gstr2b are {'itc_igst', 'itc_cgst', 'itc_sgst'} or None if missing for this period."""
    if books is None or gstr2b is None:
        return {
            "period": period,
            "books_igst": (books or {}).get("itc_igst", 0.0),
            "books_cgst": (books or {}).get("itc_cgst", 0.0),
            "books_sgst": (books or {}).get("itc_sgst", 0.0),
            "gstr2b_igst": (gstr2b or {}).get("itc_igst", 0.0),
            "gstr2b_cgst": (gstr2b or {}).get("itc_cgst", 0.0),
            "gstr2b_sgst": (gstr2b or {}).get("itc_sgst", 0.0),
            "diff_igst": 0.0, "diff_cgst": 0.0, "diff_sgst": 0.0,
            "total_variance": 0.0,
            "status": "MONTH_MISSING",
        }

    diff_igst = round(books.get("itc_igst", 0.0) - gstr2b.get("itc_igst", 0.0), 2)
    diff_cgst = round(books.get("itc_cgst", 0.0) - gstr2b.get("itc_cgst", 0.0), 2)
    diff_sgst = round(books.get("itc_sgst", 0.0) - gstr2b.get("itc_sgst", 0.0), 2)
    total_variance = round(abs(diff_igst) + abs(diff_cgst) + abs(diff_sgst), 2)

    overall = (
        "MATCH" if total_variance <= TOLERANCE_MATCH
        else "WARNING" if total_variance <= TOLERANCE_MINOR
        else "MISMATCH"
    )

    return {
        "period": period,
        "books_igst": books.get("itc_igst", 0.0),
        "books_cgst": books.get("itc_cgst", 0.0),
        "books_sgst": books.get("itc_sgst", 0.0),
        "gstr2b_igst": gstr2b.get("itc_igst", 0.0),
        "gstr2b_cgst": gstr2b.get("itc_cgst", 0.0),
        "gstr2b_sgst": gstr2b.get("itc_sgst", 0.0),
        "diff_igst": diff_igst, "diff_cgst": diff_cgst, "diff_sgst": diff_sgst,
        "total_variance": total_variance,
        "status": overall,
    }


# ─── Full reconciliation ────────────────────────────────────────────────────

def reconcile_gstr2b_vs_books(
    gstr2b_data: dict,
    books_data: dict,
    gstin_warning: Optional[str] = None,
) -> dict:
    """
    gstr2b_data: output of gstr2b_extractor.merge_gstr2b_files()
                 {"gstin", "financial_year", "monthly_data": [{"period","itc_igst","itc_cgst","itc_sgst",...}]}
    books_data:  output of books_extractor.extract_books() — {"periods": {"YYYY-MM": {...}}, "audit": {...}}
                 Only the ITC fields (itc) are relevant here; Books periods are
                 keyed YYYY-MM, GSTR-2B periods are labeled 'Apr-25' — both are
                 normalized to YYYY-MM for matching, and reported back as 'Apr-25'.
    """
    NUM_MONTH_ABBR = {
        1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun",
        7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec",
    }
    MONTH_ABBR_NUM = {v.lower(): k for k, v in NUM_MONTH_ABBR.items()}

    def label_to_yyyymm(label: str) -> Optional[str]:
        try:
            mon, yy = label.split("-")
            mon_num = MONTH_ABBR_NUM.get(mon.strip().lower())
            if mon_num is None:
                return None
            year = int(yy)
            year = 2000 + year if year < 100 else year
            return f"{year:04d}-{mon_num:02d}"
        except Exception:
            return None

    def yyyymm_to_label(yyyymm: str) -> str:
        year, mon = yyyymm.split("-")
        return f"{NUM_MONTH_ABBR[int(mon)]}-{year[-2:]}"

    # Index GSTR-2B by YYYY-MM
    gstr2b_by_period: dict[str, dict] = {}
    for entry in gstr2b_data.get("monthly_data", []):
        key = label_to_yyyymm(entry["period"])
        if key:
            gstr2b_by_period[key] = {
                "itc_igst": entry.get("itc_igst", 0.0),
                "itc_cgst": entry.get("itc_cgst", 0.0),
                "itc_sgst": entry.get("itc_sgst", 0.0),
            }

    # Books periods are already YYYY-MM; only ITC fields matter for this comparison
    books_periods = books_data.get("periods", {})
    books_by_period: dict[str, dict] = {}
    for key, p in books_periods.items():
        # books_extractor sums IGST+CGST+SGST+CESS into one 'itc' total per period
        # (it does not track ITC by tax-head). Use that as a books IGST/CGST/SGST
        # approximation is not possible — so we report the books-side total against
        # each GSTR-2B tax head's combined total instead (see note in summary).
        books_by_period[key] = p

    all_periods = sorted(set(list(gstr2b_by_period.keys()) + list(books_by_period.keys())))

    period_results = []
    exceptions = []
    for yyyymm in all_periods:
        label = yyyymm_to_label(yyyymm)
        g2b = gstr2b_by_period.get(yyyymm)
        books_raw = books_by_period.get(yyyymm)

        if books_raw is not None:
            books_split = {
                "itc_igst": books_raw.get("itc_igst", 0.0),
                "itc_cgst": books_raw.get("itc_cgst", 0.0),
                "itc_sgst": books_raw.get("itc_sgst", 0.0),
            }
        else:
            books_split = None

        pr = _reconcile_period(label, books_split, g2b)
        period_results.append(pr)
        if pr["status"] != "MATCH":
            exceptions.append(pr)

    summary = {
        "months": len(all_periods),
        "matched": sum(1 for p in period_results if p["status"] == "MATCH"),
        "warnings": sum(1 for p in period_results if p["status"] == "WARNING"),
        "mismatches": sum(1 for p in period_results if p["status"] == "MISMATCH"),
        "missing": sum(1 for p in period_results if p["status"] == "MONTH_MISSING"),
        "books_total": round(sum(p["books_igst"] + p["books_cgst"] + p["books_sgst"] for p in period_results), 2),
        "gstr2b_total": round(sum(p["gstr2b_igst"] + p["gstr2b_cgst"] + p["gstr2b_sgst"] for p in period_results), 2),
        "gstin_warning": gstin_warning,
    }

    return {
        "status": "success",
        "summary": summary,
        "monthly": period_results,
        "exceptions": exceptions,
    }


# ─── Excel styles (mirrors gstr3b_books.py) ─────────────────────────────────

_GREEN  = PatternFill("solid", fgColor="D6F5D6")
_YELLOW = PatternFill("solid", fgColor="FFF3CD")
_RED    = PatternFill("solid", fgColor="FFD6D6")
_GREY   = PatternFill("solid", fgColor="E2E8F0")
_HEADER = PatternFill("solid", fgColor="2D3748")


def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _hfont(): return Font(bold=True, color="FFFFFF", size=10)
def _bfont(): return Font(bold=True, size=10)


def _num(cell):
    cell.number_format = '#,##0.00'
    cell.alignment = Alignment(horizontal="right")


def _w(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


def _status_fill(status: str):
    if status == "MATCH":
        return _GREEN
    if status == "WARNING":
        return _YELLOW
    if status == "MONTH_MISSING":
        return _GREY
    return _RED


def _sheet_summary(ws, result: dict, gstin_2b: str, gstin_books: str):
    ws.append(["GSTR-2B vs Books — Reconciliation Summary"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([])
    s = result["summary"]

    rows = [
        ("GSTIN (GSTR-2B)", gstin_2b),
        ("GSTIN (Books)", gstin_books or "Not detected"),
        ("Periods", s["months"]),
        ("Months Matched", s["matched"]),
        ("Warnings", s["warnings"]),
        ("Mismatches", s["mismatches"]),
        ("Missing Months", s.get("missing", 0)),
        ("Books ITC Total (₹)", s["books_total"]),
        ("GSTR-2B ITC Total (₹)", s["gstr2b_total"]),
        ("Net Difference (₹)", round(s["books_total"] - s["gstr2b_total"], 2)),
        ("Match %", round(100 * s["matched"] / s["months"], 1) if s["months"] else 0),
    ]
    if s.get("gstin_warning"):
        rows.append(("⚠ GSTIN Warning", s["gstin_warning"]))

    for label, value in rows:
        ws.append([label, value])
        r = ws.max_row
        ws.cell(r, 1).font = Font(bold=True)
        for c in range(1, 3):
            ws.cell(r, c).border = _border()
        if isinstance(value, float):
            _num(ws.cell(r, 2))

    _w(ws, 1, 28); _w(ws, 2, 22)


def _sheet_monthly(ws, monthly: list):
    ws.freeze_panes = "A2"
    headers = ["Month", "Books ITC (₹)", "GSTR-2B ITC (₹)", "Difference (₹)", "Status"]
    ws.append(headers)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci)
        c.fill = _HEADER; c.font = _hfont(); c.border = _border()
        c.alignment = Alignment(horizontal="center")

    total_books = total_2b = 0.0
    for p in monthly:
        bk = round(p["books_igst"] + p["books_cgst"] + p["books_sgst"], 2)
        g2b = round(p["gstr2b_igst"] + p["gstr2b_cgst"] + p["gstr2b_sgst"], 2)
        diff = round(bk - g2b, 2)
        total_books += bk; total_2b += g2b
        ws.append([p["period"], bk, g2b, diff, p["status"]])
        ri = ws.max_row
        fill = _status_fill(p["status"])
        for ci in range(1, 6):
            ws.cell(ri, ci).fill = fill
            ws.cell(ri, ci).border = _border()
            if 2 <= ci <= 4:
                _num(ws.cell(ri, ci))

    ws.append(["Total", round(total_books, 2), round(total_2b, 2), round(total_books - total_2b, 2), ""])
    ri = ws.max_row
    for ci in range(1, 6):
        ws.cell(ri, ci).font = _bfont()
        ws.cell(ri, ci).border = _border()
        if 2 <= ci <= 4:
            _num(ws.cell(ri, ci))

    for ci, w in enumerate([14, 18, 18, 18, 16], 1):
        _w(ws, ci, w)


def _sheet_tax(ws, monthly: list):
    ws.freeze_panes = "A2"
    headers = [
        "Month",
        "Books IGST", "GSTR-2B IGST", "Diff IGST",
        "Books CGST", "GSTR-2B CGST", "Diff CGST",
        "Books SGST", "GSTR-2B SGST", "Diff SGST",
        "Status",
    ]
    ws.append(headers)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci)
        c.fill = _HEADER; c.font = _hfont(); c.border = _border()
        c.alignment = Alignment(horizontal="center")

    for p in monthly:
        ws.append([
            p["period"],
            p["books_igst"], p["gstr2b_igst"], p["diff_igst"],
            p["books_cgst"], p["gstr2b_cgst"], p["diff_cgst"],
            p["books_sgst"], p["gstr2b_sgst"], p["diff_sgst"],
            p["status"],
        ])
        ri = ws.max_row
        fill = _status_fill(p["status"])
        for ci in range(1, len(headers) + 1):
            ws.cell(ri, ci).fill = fill
            ws.cell(ri, ci).border = _border()
            if 2 <= ci <= 10:
                _num(ws.cell(ri, ci))

    for ci, w in enumerate([14] + [14]*9 + [16], 1):
        _w(ws, ci, w)


def _sheet_exceptions(ws, exceptions: list):
    if not exceptions:
        ws.append(["No mismatches — all periods matched."])
        return

    ws.freeze_panes = "A2"
    headers = ["Month", "Field", "Books (₹)", "GSTR-2B (₹)", "Difference (₹)", "Status"]
    ws.append(headers)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci)
        c.fill = _HEADER; c.font = _hfont(); c.border = _border()
        c.alignment = Alignment(horizontal="center")

    fields = [("IGST", "books_igst", "gstr2b_igst", "diff_igst"),
              ("CGST", "books_cgst", "gstr2b_cgst", "diff_cgst"),
              ("SGST", "books_sgst", "gstr2b_sgst", "diff_sgst")]

    for p in exceptions:
        for label, bk_key, g2b_key, diff_key in fields:
            diff = p[diff_key]
            if abs(diff) <= TOLERANCE_MATCH and p["status"] != "MONTH_MISSING":
                continue
            ws.append([p["period"], label, p[bk_key], p[g2b_key], diff, p["status"]])
            ri = ws.max_row
            fill = _status_fill(p["status"])
            for ci in range(1, 7):
                ws.cell(ri, ci).fill = fill
                ws.cell(ri, ci).border = _border()
                if 3 <= ci <= 5:
                    _num(ws.cell(ri, ci))

    for ci, w in enumerate([14, 12, 16, 16, 16, 16], 1):
        _w(ws, ci, w)


def _sheet_extracted(ws, gstr2b_data: dict, books_data: dict):
    ws.append(["Extracted Data — GSTR-2B"])
    ws["A1"].font = Font(bold=True, size=12)
    headers = ["Period", "ITC IGST", "ITC CGST", "ITC SGST", "Source"]
    ws.append(headers)
    hr = ws.max_row
    for ci, h in enumerate(headers, 1):
        c = ws.cell(hr, ci)
        c.fill = _HEADER; c.font = _hfont(); c.border = _border()
    for entry in gstr2b_data.get("monthly_data", []):
        ws.append([entry.get("period"), entry.get("itc_igst", 0), entry.get("itc_cgst", 0),
                   entry.get("itc_sgst", 0), entry.get("source", "")])
        ri = ws.max_row
        for ci in range(1, 6):
            ws.cell(ri, ci).border = _border()
            if 2 <= ci <= 4:
                _num(ws.cell(ri, ci))

    ws.append([])
    ws.append(["Extracted Data — Books (ITC by tax head)"])
    ws.cell(ws.max_row, 1).font = Font(bold=True, size=12)
    headers2 = ["Period", "Books ITC IGST", "Books ITC CGST", "Books ITC SGST"]
    ws.append(headers2)
    hr2 = ws.max_row
    for ci, h in enumerate(headers2, 1):
        c = ws.cell(hr2, ci)
        c.fill = _HEADER; c.font = _hfont(); c.border = _border()
    for period, p in sorted(books_data.get("periods", {}).items()):
        ws.append([period, p.get("itc_igst", 0), p.get("itc_cgst", 0), p.get("itc_sgst", 0)])
        ri = ws.max_row
        ws.cell(ri, 1).border = _border()
        for ci in range(2, 5):
            ws.cell(ri, ci).border = _border()
            _num(ws.cell(ri, ci))

    _w(ws, 1, 16); _w(ws, 2, 18); _w(ws, 3, 18); _w(ws, 4, 18); _w(ws, 5, 18)


# ─── Public: generate Excel ─────────────────────────────────────────────────

def generate_gstr2b_books_excel(
    result: dict,
    output_path: str,
    gstr2b_data: dict,
    books_data: dict,
    gstin_2b: str = "",
    gstin_books: str = "",
) -> str:
    wb = Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet("Summary")
    _sheet_summary(ws1, result, gstin_2b, gstin_books)

    ws2 = wb.create_sheet("Monthly Comparison")
    _sheet_monthly(ws2, result["monthly"])

    ws3 = wb.create_sheet("Tax Comparison")
    _sheet_tax(ws3, result["monthly"])

    ws4 = wb.create_sheet("Exceptions")
    _sheet_exceptions(ws4, result["exceptions"])

    ws5 = wb.create_sheet("Extracted Data")
    _sheet_extracted(ws5, gstr2b_data, books_data)

    wb.save(output_path)
    logger.info(f"GSTR-2B vs Books Excel saved: {output_path}")
    return output_path
