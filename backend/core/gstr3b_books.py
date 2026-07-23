import logging
import datetime
import tempfile
import os
from typing import Optional
import re

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ─── Recon row definitions ────────────────────────────────────────────────────

RECON_ROWS = [
    # Outward supplies
    ("taxable_value",        "Taxable Value"),
    ("zero_rated_value",     "Zero Rated Supplies"),
    ("nil_rated_value",      "Nil Rated Supplies"),
    ("igst_output",          "IGST Output"),
    ("cgst_output",          "CGST Output"),
    ("sgst_output",          "SGST Output"),
    # RCM inward
    ("rcm_taxable",          "RCM Taxable Value"),
    ("rcm_igst",             "RCM IGST"),
    ("rcm_cgst",             "RCM CGST"),
    ("rcm_sgst",             "RCM SGST"),
    # ITC available
    ("itc_igst",             "ITC Available IGST"),
    ("itc_cgst",             "ITC Available CGST"),
    ("itc_sgst",             "ITC Available SGST"),
    # ITC reversed
    ("itc_reversed_igst",    "ITC Reversed IGST"),
    ("itc_reversed_cgst",    "ITC Reversed CGST"),
    ("itc_reversed_sgst",    "ITC Reversed SGST"),
    # Ineligible ITC
    ("ineligible_igst",      "Ineligible ITC IGST"),
    ("ineligible_cgst",      "Ineligible ITC CGST"),
    ("ineligible_sgst",      "Ineligible ITC SGST"),
    # Net ITC
    ("net_itc_igst",         "Net ITC IGST"),
    ("net_itc_cgst",         "Net ITC CGST"),
    ("net_itc_sgst",         "Net ITC SGST"),
    # ITC utilised
    ("itc_util_igst",        "ITC Utilised IGST"),
    ("itc_util_cgst",        "ITC Utilised CGST"),
    ("itc_util_sgst",        "ITC Utilised SGST"),
    # Net tax payable / cash
    ("net_tax_payable",      "Net Tax Payable"),
    ("cash_paid_igst",       "Cash Paid IGST"),
    ("cash_paid_cgst",       "Cash Paid CGST"),
    ("cash_paid_sgst",       "Cash Paid SGST"),
    ("interest_igst",        "Interest IGST"),
    ("interest_cgst",        "Interest CGST"),
    ("interest_sgst",        "Interest SGST"),
    ("late_fee_cgst",        "Late Fee CGST"),
    ("late_fee_sgst",        "Late Fee SGST"),
]
# Groups RECON_ROWS keys into the same sections as the Books template,
# so the Excel export can mirror the UI / template layout.
DISPLAY_SECTIONS = [
    ("Tax on Outward Supply", [
        ("taxable_value", "Taxable Value"), ("zero_rated_value", "Zero Rated"),
        ("nil_rated_value", "Nil Rated"), ("igst_output", "IGST"),
        ("cgst_output", "CGST"), ("sgst_output", "SGST"),
    ]),
    ("RCM (Inward Reverse Charge)", [
        ("rcm_taxable", "Taxable Value"), ("rcm_igst", "IGST"),
        ("rcm_cgst", "CGST"), ("rcm_sgst", "SGST"),
    ]),
    ("ITC Available", [
        ("itc_igst", "IGST"), ("itc_cgst", "CGST"), ("itc_sgst", "SGST"),
    ]),
    ("ITC Reversed", [
        ("itc_reversed_igst", "IGST"), ("itc_reversed_cgst", "CGST"), ("itc_reversed_sgst", "SGST"),
    ]),
    ("Ineligible ITC", [
        ("ineligible_igst", "IGST"), ("ineligible_cgst", "CGST"), ("ineligible_sgst", "SGST"),
    ]),
    ("Net ITC Available", [
        ("net_itc_igst", "IGST"), ("net_itc_cgst", "CGST"), ("net_itc_sgst", "SGST"),
    ]),
    ("ITC Utilised", [
        ("itc_util_igst", "IGST"), ("itc_util_cgst", "CGST"), ("itc_util_sgst", "SGST"),
    ]),
    ("Net Tax Payable", [("net_tax_payable", "Total")]),
    ("Amount Paid in Cash", [
        ("cash_paid_igst", "IGST"), ("cash_paid_cgst", "CGST"), ("cash_paid_sgst", "SGST"),
    ]),
    ("Interest", [
        ("interest_igst", "IGST"), ("interest_cgst", "CGST"), ("interest_sgst", "SGST"),
    ]),
    ("Late Fee", [
        ("late_fee_cgst", "CGST"), ("late_fee_sgst", "SGST"),
    ]),
]
TOLERANCE_MATCH  = 1.0
TOLERANCE_MINOR  = 100.0

def _status(diff: float) -> str:
    abs_diff = abs(diff)
    if abs_diff <= TOLERANCE_MATCH:  return "Match"
    if abs_diff <= TOLERANCE_MINOR:  return "Difference"
    return "Mismatch"

# ─── GSTR-3B normalizer ───────────────────────────────────────────────────────

def normalize_gstr3b(g3b: dict) -> dict:
    def sf(v):
        try:
            return round(float(str(v).replace(",", "") or 0), 2)
        except Exception:
            return 0.0

    return {
        # Outward
        "taxable_value":     sf(g3b.get("outward_taxable_value", 0)),
        "zero_rated_value":  sf(g3b.get("zero_rated_value", 0)),
        "nil_rated_value":   sf(g3b.get("nil_exempt_value", 0)),
        "igst_output":       sf(g3b.get("outward_igst", 0)),
        "cgst_output":       sf(g3b.get("outward_cgst", 0)),
        "sgst_output":       sf(g3b.get("outward_sgst", 0)),
        # RCM
        "rcm_taxable":       sf(g3b.get("inward_reverse_charge_value", 0)),
        "rcm_igst":          sf(g3b.get("inward_reverse_charge_igst", 0)),
        "rcm_cgst":          sf(g3b.get("inward_reverse_charge_cgst", 0)),
        "rcm_sgst":          sf(g3b.get("inward_reverse_charge_sgst", 0)),
        # ITC available
        "itc_igst":          sf(g3b.get("other_itc_igst", 0)) + sf(g3b.get("import_goods_igst", 0)) + sf(g3b.get("import_services_igst", 0)) + sf(g3b.get("isd_igst", 0)),
        "itc_cgst":          sf(g3b.get("other_itc_cgst", 0)) + sf(g3b.get("import_goods_cgst", 0)) + sf(g3b.get("import_services_cgst", 0)) + sf(g3b.get("isd_cgst", 0)),
        "itc_sgst":          sf(g3b.get("other_itc_sgst", 0)) + sf(g3b.get("import_goods_sgst", 0)) + sf(g3b.get("import_services_sgst", 0)) + sf(g3b.get("isd_sgst", 0)),
        # ITC reversed
        "itc_reversed_igst": sf(g3b.get("rules_igst", 0)) + sf(g3b.get("others_igst", 0)),
        "itc_reversed_cgst": sf(g3b.get("rules_cgst", 0)) + sf(g3b.get("others_cgst", 0)),
        "itc_reversed_sgst": sf(g3b.get("rules_sgst", 0)) + sf(g3b.get("others_sgst", 0)),
        # Ineligible ITC
        "ineligible_igst":   sf(g3b.get("ineligible_itc_igst", 0)),
        "ineligible_cgst":   sf(g3b.get("ineligible_itc_cgst", 0)),
        "ineligible_sgst":   sf(g3b.get("ineligible_itc_sgst", 0)),
        # Net ITC
        "net_itc_igst":      sf(g3b.get("net_itc_igst", 0)),
        "net_itc_cgst":      sf(g3b.get("net_itc_cgst", 0)),
        "net_itc_sgst":      sf(g3b.get("net_itc_sgst", 0)),
        # ITC utilised (from payment table)
        "itc_util_igst":     sf(g3b.get("igst_paid_itc_igst", 0)),
        "itc_util_cgst":     sf(g3b.get("cgst_paid_itc_cgst", 0)),
        "itc_util_sgst":     sf(g3b.get("sgst_paid_itc_sgst", 0)),
        # Net tax payable and cash paid
        "net_tax_payable":   sf(g3b.get("igst_net_payable", 0)) + sf(g3b.get("cgst_net_payable", 0)) + sf(g3b.get("sgst_net_payable", 0)),
        "cash_paid_igst":    sf(g3b.get("igst_paid_cash", 0)),
        "cash_paid_cgst":    sf(g3b.get("cgst_paid_cash", 0)),
        "cash_paid_sgst":    sf(g3b.get("sgst_paid_cash", 0)),
        "interest_igst":     sf(g3b.get("igst_interest", 0)),
        "interest_cgst":     sf(g3b.get("cgst_interest", 0)),
        "interest_sgst":     sf(g3b.get("sgst_interest", 0)),
        "late_fee_cgst":     sf(g3b.get("cgst_late_fee", 0)),
        "late_fee_sgst":     sf(g3b.get("sgst_late_fee", 0)),
    }
# ─── Per-period reconciliation ────────────────────────────────────────────────

def reconcile_period(period: str, books: dict, g3b_norm: dict) -> dict:
    rows = []
    total_variance = 0.0
    for key, description in RECON_ROWS:
        bv   = books.get(key, 0.0)
        g3bv = g3b_norm.get(key, 0.0)
        diff = round(bv - g3bv, 2)
        total_variance += abs(diff)
        rows.append({
            "description": description,
            "books":       bv,
            "gstr3b":      g3bv,
            "difference":  diff,
            "status":      _status(diff),
        })
    overall = (
        "Matched" if total_variance <= TOLERANCE_MATCH
        else "Difference" if total_variance <= TOLERANCE_MINOR
        else "Mismatch"
    )
    return {
        "period":         period,
        "rows":           rows,
        "total_variance": round(total_variance, 2),
        "overall_status": overall,
    }

# ─── Full reconciliation ──────────────────────────────────────────────────────

def reconcile_gstr3b_vs_books(
    g3b_records: list[dict],
    books_data:  dict,
    gstin_warning: Optional[str] = None,
) -> dict:
    MONTH_NUM = {
        "april": 4, "may": 5, "june": 6, "july": 7, "august": 8, "september": 9,
        "october": 10, "november": 11, "december": 12,
        "january": 1, "february": 2, "march": 3,
    }

    def get_period(r: dict) -> Optional[str]:
        """Normalize GSTR-3B period to YYYY-MM to match Books period keys."""
        tp = str(r.get("TaxPeriod") or r.get("tax_period") or r.get("period") or "").strip()
        fy = str(r.get("FinancialYear") or r.get("year") or "").strip()

        month_lc = tp.lower()
        month_num = None
        for name, num in MONTH_NUM.items():
            if name in month_lc:
                month_num = num
                break

        if month_num is None:
            return None  # cannot normalize — will be skipped/logged

        # FY like "2025-26" -> first year 2025, second year 2026
        fy_match = re.match(r"(\d{4})", fy)
        first_year = int(fy_match.group(1)) if fy_match else None

        if first_year is None:
            return None

        # April-December belong to the first year of FY; Jan-March belong to the second year
        calendar_year = first_year if month_num >= 4 else first_year + 1
        return f"{calendar_year}-{month_num:02d}"

    g3b_by_period = {}
    for r in g3b_records:
        p = get_period(r)
        if p:
            g3b_by_period[p] = normalize_gstr3b(r)

    books_periods = books_data.get("periods", {})

    # Union of all periods
    all_periods = sorted(
        set(list(g3b_by_period.keys()) + list(books_periods.keys()))
    )

    period_results = []
    exceptions     = []
    for period in all_periods:
        books_p = books_periods.get(period, {
            k: 0.0 for k, _ in RECON_ROWS
        })
        g3b_p = g3b_by_period.get(period, {
            k: 0.0 for k, _ in RECON_ROWS
        })
        pr = reconcile_period(period, books_p, g3b_p)
        period_results.append(pr)
        if pr["overall_status"] != "Matched":
            exceptions.append(pr)

    # Summary totals
    summary = {
        "books_total": round(sum(
            sum(books_periods.get(p, {}).get(k, 0) for k, _ in RECON_ROWS)
            for p in all_periods
        ), 2),
        "gstr3b_total": round(sum(
            sum(g3b_by_period.get(p, {}).get(k, 0) for k, _ in RECON_ROWS)
            for p in all_periods
        ), 2),
        "months": len(all_periods),
        "matched": sum(1 for p in period_results if p["overall_status"] == "Matched"),
        "mismatches": sum(1 for p in period_results if p["overall_status"] == "Mismatch"),
        "minor": sum(1 for p in period_results if p["overall_status"] == "Difference"),
        "gstin_warning": gstin_warning,
        "audit": books_data.get("audit", {}),
    }

    return {
        "status":   "success",
        "summary":  summary,
        "monthly":  period_results,
        "exceptions": exceptions,
    }

# ─── Excel styles ─────────────────────────────────────────────────────────────

_GREEN  = PatternFill("solid", fgColor="D6F5D6")
_YELLOW = PatternFill("solid", fgColor="FFF3CD")
_RED    = PatternFill("solid", fgColor="FFD6D6")
_HEADER = PatternFill("solid", fgColor="2D3748")
_SUB    = PatternFill("solid", fgColor="F0F4F8")

def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _hfont(): return Font(bold=True, color="FFFFFF", size=10)
def _bfont(): return Font(bold=True, size=10)

def _num(cell):
    cell.number_format = '#,##0.00'
    cell.alignment = Alignment(horizontal="right")

def _w(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def _status_fill(status: str):
    if status == "Matched":  return _GREEN
    if status == "Difference": return _YELLOW
    return _RED

# ─── Sheet writers ────────────────────────────────────────────────────────────

def _sheet_summary(ws, result: dict, gstin_3b: str, gstin_books: str):
    ws.append(["GSTR-3B vs Books — Reconciliation Summary"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([])
    s = result["summary"]

    rows = [
        ("GSTIN (GSTR-3B)", gstin_3b),
        ("GSTIN (Books)",   gstin_books or "Not detected"),
        ("Periods",         s["months"]),
        ("Months Matched",  s["matched"]),
        ("Minor Differences", s["minor"]),
        ("Mismatches",      s["mismatches"]),
        ("Books Total (₹)", s["books_total"]),
        ("GSTR-3B Total (₹)", s["gstr3b_total"]),
        ("Net Difference (₹)", round(s["books_total"] - s["gstr3b_total"], 2)),
    ]
    if s.get("gstin_warning"):
        rows.append(("⚠ GSTIN Warning", s["gstin_warning"]))

    for label, value in rows:
        ws.append([label, value])
        r = ws.max_row
        ws.cell(r, 1).font = Font(bold=True)
        for c in range(1, 3):
            ws.cell(r, c).border = _border()
        if isinstance(value, float):
            _num(ws.cell(r, 2))

    _w(ws, 1, 30); _w(ws, 2, 22)


def _sheet_monthly(ws, monthly: list):
    ws.freeze_panes = "A3"
    headers = ["Period", "Books (₹)", "GSTR-3B (₹)", "Difference (₹)", "Status"]
    ws.append(headers)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci)
        c.fill = _HEADER; c.font = _hfont(); c.border = _border()
        c.alignment = Alignment(horizontal="center")

    # One row per period (summed across all RECON_ROWS)
    for p in monthly:
        bk  = sum(r["books"]   for r in p["rows"])
        g3b = sum(r["gstr3b"]  for r in p["rows"])
        diff = round(bk - g3b, 2)
        ws.append([p["period"], bk, g3b, diff, p["overall_status"]])
        ri = ws.max_row
        fill = _status_fill(p["overall_status"])
        for ci in range(1, 6):
            ws.cell(ri, ci).fill = fill
            ws.cell(ri, ci).border = _border()
            if 2 <= ci <= 4:
                _num(ws.cell(ri, ci))

    for ci, w in enumerate([16, 18, 18, 18, 18], 1):
        _w(ws, ci, w)


def _sheet_tax(ws, monthly: list):
    ws.freeze_panes = "A3"
    fields = ["igst", "cgst", "sgst", "cess", "itc", "rcm"]
    headers = ["Period"] + [f.upper() for f in fields] + ["Status"]
    ws.append(headers)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci)
        c.fill = _HEADER; c.font = _hfont(); c.border = _border()
        c.alignment = Alignment(horizontal="center")

    desc_map = {r["description"]: r for rows in [p["rows"] for p in monthly] for r in rows}
    for p in monthly:
        row_by_desc = {r["description"]: r for r in p["rows"]}
        label_map = {
            "igst": "IGST", "cgst": "CGST", "sgst": "SGST",
            "cess": "Cess", "itc": "ITC (Input Tax Credit)", "rcm": "RCM (Reverse Charge)"
        }
        row_data = [p["period"]]
        for f in fields:
            lbl = label_map[f]
            r = row_by_desc.get(lbl, {})
            row_data.append(r.get("difference", 0.0))
        row_data.append(p["overall_status"])
        ws.append(row_data)
        ri = ws.max_row
        fill = _status_fill(p["overall_status"])
        for ci in range(1, len(headers) + 1):
            ws.cell(ri, ci).fill = fill
            ws.cell(ri, ci).border = _border()
            if 2 <= ci <= len(fields) + 1:
                _num(ws.cell(ri, ci))

    for ci, w in enumerate([16] + [14]*len(fields) + [18], 1):
        _w(ws, ci, w)

def _sheet_books_format(ws, monthly: list):
    """Section-based layout matching the Books template / UI — each field
    gets 3 sub-columns: Books, GSTR-3B, Difference."""
    ws.freeze_panes = "B4"

    # Row 1: section headers, Row 2: field headers, Row 3: Books/3B/Diff
    ws.cell(1, 1, "Month")
    ws.merge_cells(start_row=1, start_column=1, end_row=3, end_column=1)

    col = 2
    for section_label, fields in DISPLAY_SECTIONS:
        sec_start = col
        for field_key, field_label in fields:
            field_start = col
            ws.cell(2, col, field_label)
            ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 2)
            ws.cell(3, col, "Books"); ws.cell(3, col + 1, "GSTR-3B"); ws.cell(3, col + 2, "Diff")
            col += 3
        ws.cell(1, sec_start, section_label)
        ws.merge_cells(start_row=1, start_column=sec_start, end_row=1, end_column=col - 1)

    total_cols = col - 1
    for r in (1, 2, 3):
        for ci in range(1, total_cols + 1):
            c = ws.cell(r, ci)
            c.font = _hfont()
            c.fill = _HEADER
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = _border()

    row = 4
    for p in monthly:
        row_by_key = {r["description"]: r for r in p["rows"]}
        # RECON_ROWS description text is the lookup key used in p["rows"]
        desc_by_field_key = {key: desc for key, desc in RECON_ROWS}
        ws.cell(row, 1, p["period"])
        col = 2
        for section_label, fields in DISPLAY_SECTIONS:
            for field_key, field_label in fields:
                desc = desc_by_field_key.get(field_key)
                r = row_by_key.get(desc, {})
                bv = r.get("books", 0.0); gv = r.get("gstr3b", 0.0); dv = r.get("difference", 0.0)
                status = r.get("status", "Match")
                fill = _status_fill(status)
                ws.cell(row, col, bv); ws.cell(row, col + 1, gv); ws.cell(row, col + 2, dv)
                for cc in range(col, col + 3):
                    ws.cell(row, cc).fill = fill
                    ws.cell(row, cc).border = _border()
                    _num(ws.cell(row, cc))
                col += 3
        ws.cell(row, 1).border = _border()
        row += 1

    ws.column_dimensions["A"].width = 12
    for ci in range(2, total_cols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 11
def _sheet_exceptions(ws, exceptions: list):
    if not exceptions:
        ws.append(["No mismatches — all periods matched."])
        return

    ws.freeze_panes = "A3"
    headers = ["Period", "Description", "Books (₹)", "GSTR-3B (₹)", "Difference (₹)", "Status"]
    ws.append(headers)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci)
        c.fill = _HEADER; c.font = _hfont(); c.border = _border()
        c.alignment = Alignment(horizontal="center")

    for p in exceptions:
        for row in p["rows"]:
            if row["status"] == "Match":
                continue
            ws.append([p["period"], row["description"], row["books"], row["gstr3b"], row["difference"], row["status"]])
            ri = ws.max_row
            fill = _status_fill(row["status"])
            for ci in range(1, 7):
                ws.cell(ri, ci).fill = fill
                ws.cell(ri, ci).border = _border()
                if 3 <= ci <= 5:
                    _num(ws.cell(ri, ci))

    for ci, w in enumerate([16, 28, 16, 16, 16, 18], 1):
        _w(ws, ci, w)


def _sheet_audit(ws, audit: dict):
    ws.append(["Extraction Audit"])
    ws["A1"].font = Font(bold=True, size=12)
    ws.append([])

    ws.append(["Sheets Detected"])
    ws.cell(ws.max_row, 1).font = _bfont()
    for s in audit.get("sheets_detected", []):
        ws.append([s.get("sheet",""), s.get("type","unknown")])

    ws.append([])
    ws.append(["Skipped Sheets"])
    ws.cell(ws.max_row, 1).font = _bfont()
    for s in audit.get("skipped_sheets", []):
        ws.append([s])

    ws.append([])
    ws.append(["Columns Detected per Sheet"])
    ws.cell(ws.max_row, 1).font = _bfont()
    for sheet, cols in audit.get("columns_detected", {}).items():
        ws.append([sheet])
        for field, col in cols.items():
            ws.append(["", field, col])

    _w(ws, 1, 28); _w(ws, 2, 22); _w(ws, 3, 22)


# ─── Public: generate Excel ───────────────────────────────────────────────────

def generate_gstr3b_books_excel(
    result:       dict,
    output_path:  str,
    gstin_3b:     str = "",
    gstin_books:  str = "",
) -> str:
    wb = Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet("Summary")
    _sheet_summary(ws1, result, gstin_3b, gstin_books)

    ws2 = wb.create_sheet("Monthly Comparison")
    _sheet_monthly(ws2, result["monthly"])

    ws3 = wb.create_sheet("Tax Comparison")
    _sheet_tax(ws3, result["monthly"])
    ws2b = wb.create_sheet("Books Format")
    _sheet_books_format(ws2b, result["monthly"])

    ws4 = wb.create_sheet("Exceptions")
    _sheet_exceptions(ws4, result["exceptions"])

    ws5 = wb.create_sheet("Extraction Audit")
    _sheet_audit(ws5, result["summary"].get("audit", {}))

    wb.save(output_path)
    logger.info(f"GSTR-3B vs Books Excel saved: {output_path}")
    return output_path