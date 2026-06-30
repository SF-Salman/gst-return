"""
GSTR-2B extractor (Excel format).

This is distinct from backend/core/extractors/gstr2b.py, which parses the
GSTN portal's *JSON* download into InvoiceRecord line items. This module
instead parses the GSTN portal's *Excel* download (the .xlsx a taxpayer
downloads from the GST portal — 'Read me', 'ITC Available', 'B2B', ... sheets)
and extracts the period-level ITC summary needed for GSTR-2B vs Books
reconciliation.

Preferred source: the 'ITC Available' sheet's "FORM SUMMARY" block, which is
the same headline figures the portal itself reports against GSTR-3B Table
4(A) — far more reliable than re-deriving totals from individual B2B line
items, and is what a reviewing accountant would tie out against anyway.

Fallback: if 'ITC Available' isn't present/parseable, sum tax columns
directly off the 'B2B' invoice-level sheet.
"""

import logging
import re
from typing import Optional

import openpyxl

logger = logging.getLogger(__name__)

MONTH_NUM = {
    "january": 1, "february": 2, "march": 3, "april": 4, "may": 5, "june": 6,
    "july": 7, "august": 8, "september": 9, "october": 10, "november": 11, "december": 12,
}
NUM_MONTH_ABBR = {
    1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun",
    7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec",
}


def _safe_float(v) -> float:
    try:
        if v is None or v == "":
            return 0.0
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).replace(",", "").strip()
        return float(s) if s else 0.0
    except (ValueError, TypeError):
        return 0.0


def _period_to_label(financial_year: str, tax_period: str) -> Optional[str]:
    """Build 'Apr-25' style label from Read me sheet's 'Financial Year' (e.g. '2025-26')
    and 'Tax Period' (e.g. 'April')."""
    month_num = MONTH_NUM.get((tax_period or "").strip().lower())
    if month_num is None:
        return None
    fy_match = re.match(r"(\d{4})", str(financial_year or "").strip())
    if not fy_match:
        return None
    first_year = int(fy_match.group(1))
    # April–December belong to the first year of FY; Jan–March belong to the second year
    calendar_year = first_year if month_num >= 4 else first_year + 1
    return f"{NUM_MONTH_ABBR[month_num]}-{str(calendar_year)[-2:]}"


def _read_header_info(wb) -> dict:
    """Pull GSTIN / Financial Year / Tax Period from the 'Read me' sheet, which is
    present on every official GSTR-2B Excel download and uses a stable label/value
    layout regardless of which columns the rest of the sheet uses."""
    info = {"gstin": "", "financial_year": "", "tax_period": "", "legal_name": ""}
    if "Read me" not in wb.sheetnames:
        return info

    ws = wb["Read me"]
    for row in ws.iter_rows(min_row=1, max_row=15):
        cells = [c.value for c in row]
        label = str(cells[0]).strip().lower() if cells and cells[0] else ""
        value = next((c for c in cells[1:] if c not in (None, "")), None)
        if label == "financial year":
            info["financial_year"] = str(value or "").strip()
        elif label == "tax period":
            info["tax_period"] = str(value or "").strip()
        elif label == "gstin":
            info["gstin"] = str(value or "").strip()
        elif label == "legal name":
            info["legal_name"] = str(value or "").strip()
    return info


def _extract_from_itc_available(wb) -> Optional[dict]:
    """Sum the official ITC-available headline rows (the GSTR-3B-table-referenced
    rows in column C — 4(A)(1), 4(A)(3), 4(A)(4), 4(A)(5)), net of Part B credit
    notes (4(A) row under Part B). Returns None if the sheet/expected layout isn't found."""
    if "ITC Available" not in wb.sheetnames:
        return None
    ws = wb["ITC Available"]

    igst = cgst = sgst = cess = 0.0
    found_any = False

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        cells = [c.value for c in row]
        if len(cells) < 7:
            continue
        gstr3b_table = str(cells[2] or "").strip()
        # Headline rows reference a GSTR-3B table like '4(A)(5)', '4(A)(4)', '4(A)(1)'.
        # Skip 'Details' sub-rows (no table ref) and RCM-liability rows (3.1(d) / 4(A)(3)
        # combined ref — RCM ITC is tracked, but RCM liability itself is not an ITC figure).
        if not gstr3b_table or "Details" in str(cells[0] or ""):
            continue
        is_credit_note_section = "4(a)" == gstr3b_table.lower() and not re.search(r"\(\d\)\(\d\)", gstr3b_table)

        row_igst = _safe_float(cells[3])
        row_cgst = _safe_float(cells[4])
        row_sgst = _safe_float(cells[5])
        row_cess = _safe_float(cells[6])

        if is_credit_note_section:
            # Part B — credit notes net off against ITC available
            igst -= row_igst; cgst -= row_cgst; sgst -= row_sgst; cess -= row_cess
            found_any = True
        elif re.match(r"^4\(a\)\([1345]\)$", gstr3b_table.lower()):
            igst += row_igst; cgst += row_cgst; sgst += row_sgst; cess += row_cess
            found_any = True

    if not found_any:
        return None

    return {
        "itc_igst": round(igst, 2),
        "itc_cgst": round(cgst, 2),
        "itc_sgst": round(sgst, 2),
        "itc_cess": round(cess, 2),
    }


def _extract_from_b2b_fallback(wb) -> Optional[dict]:
    """Fallback: sum IGST/CGST/SGST columns directly from the 'B2B' invoice sheet
    if the summary sheet isn't usable. Less reliable (excludes ISD/import/RCM ITC,
    doesn't net credit notes) but better than nothing."""
    if "B2B" not in wb.sheetnames:
        return None
    ws = wb["B2B"]

    header_row_idx = None
    header_map = {}
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15), start=1):
        cells = [str(c.value).strip().lower() if c.value else "" for c in row]
        if any("integrated tax" in c for c in cells) and any("central tax" in c for c in cells):
            header_row_idx = idx
            for ci, c in enumerate(cells):
                if "integrated tax" in c:
                    header_map["igst"] = ci
                elif "central tax" in c:
                    header_map["cgst"] = ci
                elif "state/ut tax" in c or "state tax" in c:
                    header_map["sgst"] = ci
                elif c == "cess" or "cess" in c:
                    header_map["cess"] = ci
            break

    if header_row_idx is None or "igst" not in header_map:
        return None

    igst = cgst = sgst = cess = 0.0
    for row in ws.iter_rows(min_row=header_row_idx + 1, max_row=ws.max_row):
        cells = [c.value for c in row]
        if len(cells) <= header_map["igst"]:
            continue
        igst += _safe_float(cells[header_map.get("igst")]) if header_map.get("igst") is not None and header_map["igst"] < len(cells) else 0.0
        cgst += _safe_float(cells[header_map.get("cgst")]) if header_map.get("cgst") is not None and header_map["cgst"] < len(cells) else 0.0
        sgst += _safe_float(cells[header_map.get("sgst")]) if header_map.get("sgst") is not None and header_map["sgst"] < len(cells) else 0.0
        cess += _safe_float(cells[header_map.get("cess")]) if header_map.get("cess") is not None and header_map["cess"] < len(cells) else 0.0

    return {
        "itc_igst": round(igst, 2),
        "itc_cgst": round(cgst, 2),
        "itc_sgst": round(sgst, 2),
        "itc_cess": round(cess, 2),
    }


def extract_gstr2b(file_bytes: bytes, filename: str = "") -> dict:
    """
    Parse one GSTR-2B Excel download (one file = one month, per portal convention)
    and return:
    {
      "gstin": "...",
      "financial_year": "2025-26",
      "monthly_data": [
        {"period": "Apr-25", "itc_igst": .., "itc_cgst": .., "itc_sgst": .., "itc_cess": .., "source": "ITC Available" | "B2B (fallback)"}
      ]
    }

    Raises ValueError if neither the summary sheet nor the B2B fallback can be parsed.
    """
    import io
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    header = _read_header_info(wb)
    period_label = _period_to_label(header["financial_year"], header["tax_period"])
    if not period_label:
        # Fall back to filename convention: MMYYYY_GSTIN_GSTR2B_*.xlsx
        m = re.match(r"^(\d{2})(\d{4})_", filename or "")
        if m:
            mm, yyyy = int(m.group(1)), m.group(2)
            if 1 <= mm <= 12:
                period_label = f"{NUM_MONTH_ABBR[mm]}-{yyyy[-2:]}"

    source = "ITC Available"
    totals = _extract_from_itc_available(wb)
    if totals is None:
        source = "B2B (fallback)"
        totals = _extract_from_b2b_fallback(wb)

    if totals is None:
        raise ValueError(
            "Could not extract ITC figures from this GSTR-2B file. "
            "Expected an 'ITC Available' summary sheet or a 'B2B' invoice sheet "
            "matching the official GST portal Excel export format."
        )

    if not period_label:
        raise ValueError(
            "Could not determine the tax period for this GSTR-2B file. "
            "Expected 'Tax Period' / 'Financial Year' on the 'Read me' sheet."
        )

    monthly_entry = {"period": period_label, "source": source, **totals}
    logger.info(f"Extracted GSTR-2B {period_label} via {source}: {totals}")

    return {
        "gstin": header["gstin"],
        "financial_year": header["financial_year"],
        "legal_name": header["legal_name"],
        "monthly_data": [monthly_entry],
    }


def merge_gstr2b_files(extracted_list: list[dict]) -> dict:
    """Combine multiple single-month extract_gstr2b() outputs (one per uploaded file)
    into one {gstin, financial_year, monthly_data: [...]} covering all months,
    sorted by period. Later files override earlier ones for the same period."""
    by_period: dict[str, dict] = {}
    gstin = ""
    financial_year = ""
    for ext in extracted_list:
        gstin = gstin or ext.get("gstin", "")
        financial_year = financial_year or ext.get("financial_year", "")
        for entry in ext.get("monthly_data", []):
            by_period[entry["period"]] = entry

    def _sort_key(period_label: str):
        mon_str, yy = period_label.split("-")
        mon_num = next((n for n, a in NUM_MONTH_ABBR.items() if a == mon_str), 0)
        year = int(yy)
        # April starts the FY ordering (Apr..Dec, then Jan..Mar)
        fy_order = mon_num - 4 if mon_num >= 4 else mon_num + 8
        return (year if mon_num >= 4 else year - 1, fy_order)

    ordered_periods = sorted(by_period.keys(), key=_sort_key)
    return {
        "gstin": gstin,
        "financial_year": financial_year,
        "monthly_data": [by_period[p] for p in ordered_periods],
    }
