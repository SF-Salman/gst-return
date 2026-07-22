import logging
import datetime
from typing import Dict, Any, List

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ─── Tolerance ────────────────────────────────────────────────────────────────

AMOUNT_TOLERANCE = 1.0


# ─── Row definitions (section, description) ──────────────────────────────────
# Each tuple: (data_key, section_ref, description)
# These are the SAME rows shown in the monthly detail table.


RECON_ROWS = [
    # 3.1(a) — Combined total (what GSTR-3B actually reports)
    ("outward_taxable_value", "3.1(a)", "Outward Taxable Value"),
    ("outward_igst",          "3.1(a)", "IGST"),
    ("outward_cgst",          "3.1(a)", "CGST"),
    ("outward_sgst",          "3.1(a)", "SGST"),
    ("outward_cess",          "3.1(a)", "Cess"),
    # 3.2 — B2CS inter-state (GSTR-3B reports this separately under 3.2)
    ("b2cs_value",            "3.2",    "B2CS Value (Inter-state)"),
    ("b2cs_igst",             "3.2",    "B2CS IGST"),
    # 9B CDNR — informational only (GSTR-3B has no counterpart, shown for audit)
    ("cdnr_value",            "9B CDNR","CDNR Adj. Value (info)"),
    ("cdnr_igst",             "9B CDNR","CDNR Adj. IGST (info)"),
    # 3.1(b) — Exports / Zero Rated
    ("zero_rated_value",      "3.1(b)", "Exports / Zero Rated Value"),
    ("zero_rated_igst",       "3.1(b)", "Exports IGST"),
    # 3.1(c) — Nil / Exempt
    ("nil_exempt_value",      "3.1(c)", "Nil / Exempt Supplies"),
    # 3.1(d) — RCM inward (GSTR-3B only — GSTR-1 4B is outward RCM, not comparable)
    ("reverse_charge_value",  "3.1(d)", "RCM Inward Value (3B only)"),
    ("reverse_charge_igst",   "3.1(d)", "RCM Inward IGST (3B only)"),
    ("reverse_charge_cgst",   "3.1(d)", "RCM Inward CGST (3B only)"),
    ("reverse_charge_sgst",   "3.1(d)", "RCM Inward SGST (3B only)"),
]

# Keys that have no GSTR-3B counterpart — shown informational only, never flagged as mismatch
INFO_ONLY_KEYS = {
    "cdnr_value", "cdnr_igst",
    "reverse_charge_value", "reverse_charge_igst",
    "reverse_charge_cgst",  "reverse_charge_sgst",
}

# Human-readable description list (used for column headers in annual sheet)
ROW_DESCRIPTIONS = [desc for _, _, desc in RECON_ROWS]

def safe_float(value) -> float:
    try:
        if isinstance(value, str):
            value = value.replace(",", "").strip()
        return round(float(value or 0), 2)
    except (ValueError, TypeError):
        return 0.0
# ─── Helpers ──────────────────────────────────────────────────────────────────

def calculate_gstr1_liability(gstr1_data: dict) -> dict:
    """
    Returns:
      {
        "values":    { ...keys matching RECON_ROWS... },
        "breakdown": { "3.1(a)": { "formula", "components", "total" }, ... }
      }
    Reverse Charge (4B) is NOT included anywhere.
    B2CL (Table 5) and B2CS (Table 7) are calculated separately
    and do NOT inflate registered_value / outward_taxable_value.
    """
    sf = safe_float

# ── Registered supplies (4A + 6B(SEZ) + 6C only) ─────────
    # ── Registered supplies (4A + 6C only) ─────────
    # NOTE: 6B (SEZ) is a zero-rated supply under GST and flows into GSTR-3B
    # 3.1(b), NOT 3.1(a). It is already included in export_value below;
    # including it here as well double-counts it in 3.1(a).
    reg_base_value = (
        sf(gstr1_data.get("4A_Value", 0)) +
        sf(gstr1_data.get("6C_Value", 0))
    )
    reg_base_igst = (
        sf(gstr1_data.get("4A_IGST", 0)) +
        sf(gstr1_data.get("6C_IGST", 0))
    )
    reg_base_cgst = (
        sf(gstr1_data.get("4A_CGST", 0)) +
        sf(gstr1_data.get("6C_CGST", 0))
    )
    reg_base_sgst = (
        sf(gstr1_data.get("4A_SGST", 0)) +
        sf(gstr1_data.get("6C_SGST", 0))
    )
    reg_base_cess = (
        sf(gstr1_data.get("4A_Cess", 0)) +
        sf(gstr1_data.get("6C_Cess", 0))
    )
    # Registered amendments (9A — net diff only, never amended totals)
    reg_amend_value = (
        sf(gstr1_data.get("9A_B2BRegular_NetDiff_Value", 0))
        + sf(gstr1_data.get("9A_SEZWP_NetDiff_Value", 0))
        + sf(gstr1_data.get("9A_SEZWOP_NetDiff_Value", 0))
        + sf(gstr1_data.get("9A_DE_NetDiff_Value", 0))
    )
    reg_amend_igst = (
        sf(gstr1_data.get("9A_B2BRegular_NetDiff_IGST",0))
        + sf(gstr1_data.get("9A_SEZWP_NetDiff_IGST",0))
        + sf(gstr1_data.get("9A_DE_NetDiff_IGST",0))
    )
    reg_amend_cgst = (
        sf(gstr1_data.get("9A_B2BRegular_NetDiff_CGST",0))
        + sf(gstr1_data.get("9A_DE_NetDiff_CGST",0))
    )
    reg_amend_sgst = (
        sf(gstr1_data.get("9A_B2BRegular_NetDiff_SGST",0))
        + sf(gstr1_data.get("9A_DE_NetDiff_SGST",0))
    )
    reg_amend_cess = (
        sf(gstr1_data.get("9A_B2BRegular_NetDiff_Cess",0))
        + sf(gstr1_data.get("9A_SEZWP_NetDiff_Cess",0))
        + sf(gstr1_data.get("9A_DE_NetDiff_Cess",0))
    )
    # 9B CDNR Registered — nets into 3.1(a)
    # Sub-categories: B2B Regular (Table 4), SEZ (Table 6B), DE (Table 6C)
    # 9B CDNR Registered — nets into 3.1(a)
    # Sub-categories: B2B Regular (Table 4), DE (Table 6C).
    # NOTE: SEZ (Table 6B) credit/debit notes are zero-rated and are already
    # netted into cdnur_export_value below — including them here as well
    # double-counts them in 3.1(a).
    cdnr_reg_value = (
        sf(gstr1_data.get("9B_CDNR_B2BRegular_Value", 0)) +
        sf(gstr1_data.get("9B_CDNR_DE_Value", 0))
    )
    cdnr_reg_igst = (
        sf(gstr1_data.get("9B_CDNR_B2BRegular_IGST", 0)) +
        sf(gstr1_data.get("9B_CDNR_DE_IGST", 0))
    )
    cdnr_reg_cgst = (
        sf(gstr1_data.get("9B_CDNR_B2BRegular_CGST", 0)) +
        sf(gstr1_data.get("9B_CDNR_DE_CGST", 0))
    )
    cdnr_reg_sgst = (
        sf(gstr1_data.get("9B_CDNR_B2BRegular_SGST", 0)) +
        sf(gstr1_data.get("9B_CDNR_DE_SGST", 0))
    )
    cdnr_reg_cess = (
        sf(gstr1_data.get("9B_CDNR_B2BRegular_Cess", 0)) +
        sf(gstr1_data.get("9B_CDNR_DE_Cess", 0))
    )

    registered_value = round(reg_base_value + reg_amend_value + cdnr_reg_value, 2)
    registered_igst  = round(reg_base_igst  + reg_amend_igst  + cdnr_reg_igst,  2)
    registered_cgst  = round(reg_base_cgst  + reg_amend_cgst  + cdnr_reg_cgst,  2)
    registered_sgst  = round(reg_base_sgst  + reg_amend_sgst  + cdnr_reg_sgst,  2)
    registered_cess  = round(reg_base_cess  + reg_amend_cess  + cdnr_reg_cess,  2)

    # ── B2CL (Table 5) — independent, NOT added to outward_taxable_value ─────
    b2cl_amend_value = sf(gstr1_data.get("9A_B2CL_NetDiff_Value", 0))
    b2cl_amend_igst  = sf(gstr1_data.get("9A_B2CL_NetDiff_IGST",  0))

    # 9B CDNUR B2CL nets into Table 5 (B2CL)
    cdnur_b2cl_value = sf(gstr1_data.get("9B_CDNUR_B2CL_Value", 0))
    cdnur_b2cl_igst  = sf(gstr1_data.get("9B_CDNUR_B2CL_IGST",  0))

    b2cl_value = round(sf(gstr1_data.get("5_Value", 0)) + b2cl_amend_value + cdnur_b2cl_value, 2)
    b2cl_igst  = round(sf(gstr1_data.get("5_IGST",  0)) + b2cl_amend_igst  + cdnur_b2cl_igst,  2)

    # ── B2CS (Table 7) — independent, NOT added to outward_taxable_value ─────
    b2cs_amend_value = sf(gstr1_data.get("10_NetDiff_Value", 0))
    b2cs_amend_igst  = sf(gstr1_data.get("10_NetDiff_IGST",  0))
    b2cs_amend_cgst  = sf(gstr1_data.get("10_NetDiff_CGST",  0))
    b2cs_amend_sgst  = sf(gstr1_data.get("10_NetDiff_SGST",  0))

    b2cs_value = round(sf(gstr1_data.get("7_Value", 0)) + b2cs_amend_value, 2)
    b2cs_igst  = round(sf(gstr1_data.get("7_IGST",  0)) + b2cs_amend_igst,  2)
    b2cs_cgst  = round(sf(gstr1_data.get("7_CGST",  0)) + b2cs_amend_cgst,  2)
    b2cs_sgst  = round(sf(gstr1_data.get("7_SGST",  0)) + b2cs_amend_sgst,  2)

   # ── 3.1(a) = registered + B2CS only (B2CL goes into 3.2, NOT 3.1(a)) ─────
    outward_taxable_value = round(registered_value + b2cs_value, 2)
    outward_igst          = round(registered_igst  + b2cs_igst,  2)
    outward_cgst          = round(registered_cgst  + b2cs_cgst,  2)
    outward_sgst          = round(registered_sgst  + b2cs_sgst,  2)
    outward_cess          = registered_cess
    # ── Amendment totals (for audit row) ─────────────────────────────────────
    amend_outward_value = round(reg_amend_value + b2cl_amend_value + b2cs_amend_value, 2)
    amend_outward_igst  = round(reg_amend_igst  + b2cl_amend_igst  + b2cs_amend_igst,  2)
    amend_outward_cgst  = round(reg_amend_cgst  + b2cs_amend_cgst,                      2)
    amend_outward_sgst  = round(reg_amend_sgst  + b2cs_amend_sgst,                      2)

    # ── 3.1(b): Exports / Zero Rated ─────────────────────────────────────────
    export_value = (sf(gstr1_data.get("6A_Value", 0))
                  + sf(gstr1_data.get("6B_Value", 0))
                  + sf(gstr1_data.get("6C_Value", 0)))
    export_igst  = (sf(gstr1_data.get("6A_IGST",  0))
                  + sf(gstr1_data.get("6B_IGST",  0))
                  + sf(gstr1_data.get("6C_IGST",  0)))

    export_amend_value = (
        sf(gstr1_data.get("9A_EXPWP_NetDiff_Value",0))
        + sf(gstr1_data.get("9A_EXPWOP_NetDiff_Value",0))
    )
    export_amend_igst = (
        sf(gstr1_data.get("9A_EXPWP_NetDiff_IGST",0))
    )

    # 9B CDNUR (EXPWP/EXPWOP) + 9B CDNR SEZ — all net into exports (6A+6B)
    cdnur_export_value = (
        sf(gstr1_data.get("9B_CDNUR_EXPWP_Value",  0))
        + sf(gstr1_data.get("9B_CDNUR_EXPWOP_Value", 0))
        + sf(gstr1_data.get("9B_CDNR_SEZ_Value",     0))
    )
    cdnur_export_igst = (
        sf(gstr1_data.get("9B_CDNUR_EXPWP_IGST",  0))
        + sf(gstr1_data.get("9B_CDNUR_EXPWOP_IGST", 0))
        + sf(gstr1_data.get("9B_CDNR_SEZ_IGST",     0))
    )

    zero_rated_value = round(export_value + export_amend_value + cdnur_export_value, 2)
    zero_rated_igst  = round(export_igst  + export_amend_igst  + cdnur_export_igst,  2)

    # ── 3.1(c): Nil / Exempt ─────────────────────────────────────────────────
    nil_exempt_value = round(sf(gstr1_data.get("8_Total", 0)), 2)

    # ── Values dict (used by reconcile_gstr1_vs_gstr3b) ──────────────────────
    values = {
        # 3.1(a) combined — matches what GSTR-3B reports in 3.1(a)
        "outward_taxable_value": outward_taxable_value,
        "outward_igst":          outward_igst,
        "outward_cgst":          outward_cgst,
        "outward_sgst":          outward_sgst,
        "outward_cess":          outward_cess,
        # 3.2 B2CS — for cross-check against GSTR-3B 3.2
        "b2cs_value":            b2cs_value,
        "b2cs_igst":             b2cs_igst,
        # 9B CDNR — informational (already netted into outward_taxable_value above)
        "cdnr_value":            round(cdnr_reg_value, 2),
        "cdnr_igst":             round(cdnr_reg_igst, 2),
        # 3.1(b)
        "zero_rated_value":      zero_rated_value,
        "zero_rated_igst":       zero_rated_igst,
        # 3.1(c)
        "nil_exempt_value":      nil_exempt_value,
        # 3.1(d) — Reverse charge (4B) — not in outward_taxable_value
        "reverse_charge_value":  round(safe_float(gstr1_data.get("4B_Value", 0)), 2),
        "reverse_charge_igst":   round(safe_float(gstr1_data.get("4B_IGST",  0)), 2),
        "reverse_charge_cgst":   round(safe_float(gstr1_data.get("4B_CGST",  0)), 2),
        "reverse_charge_sgst":   round(safe_float(gstr1_data.get("4B_SGST",  0)), 2),
        # Amendment total (kept for breakdown audit, not in RECON_ROWS)
        "amend_outward_value":   amend_outward_value,
        "amend_outward_igst":    amend_outward_igst,
        "amend_outward_cgst":    amend_outward_cgst,
        "amend_outward_sgst":    amend_outward_sgst,
    }

    # ── Breakdown dict (consumed by Excel Calculation Summary sheet + UI) ─────
    breakdown = {
        "3.1(a) Registered": {
            "formula": "4A + 6B + 6C + 9A_B2BReg/SEZ/DE(Net Diff) + 9B_CDNR(B2BReg+SEZ+DE net)",
            "components": [
                {"table": "4A",              "value": sf(gstr1_data.get("4A_Value", 0))},
                {"table": "6B",              "value": sf(gstr1_data.get("6B_Value", 0))},
                {"table": "6C",              "value": sf(gstr1_data.get("6C_Value", 0))},
                {"table": "9A Reg Net Diff", "value": reg_amend_value},
                {"table": "9B CDNR Reg",     "value": cdnr_reg_value},
            ],
            "total": registered_value,
        },
        "3.1(a) B2CL": {
            "formula": "5 + 9A_B2CL(Net Diff) + 9B_CDNUR_B2CL(net)",
            "components": [
                {"table": "5",               "value": sf(gstr1_data.get("5_Value", 0))},
                {"table": "9A B2CL Net Diff","value": b2cl_amend_value},
            ],
            "total": b2cl_value,
        },
        "3.1(a) B2CS": {
            "formula": "7 + 10(Net Diff)",
            "components": [
                {"table": "7",               "value": sf(gstr1_data.get("7_Value", 0))},
                {"table": "10 Net Diff",     "value": b2cs_amend_value},
            ],
            "total": b2cs_value,
        },
        "3.1(a) Total": {
            "formula": "Registered (4A+6B+6C+9A+9B_CDNR) + B2CS (7+10)",
            "components": [
                {"table": "Registered",      "value": registered_value},
                {"table": "B2CS (Table 7)",  "value": b2cs_value},
            ],
            "total": outward_taxable_value,
        },
        "3.1(b) Exports": {
            "formula": "6A + 9A_EXPWP/EXPWOP/SEZWP/SEZWOP(Net Diff) + 9B_CDNUR_EXPWP/EXPWOP(net)",
            "components": [
                {"table": "6A",              "value": export_value},
                {"table": "9A Exp Net Diff", "value": export_amend_value},
                {"table": "9B CDNUR Exp",    "value": cdnur_export_value},
            ],
            "total": zero_rated_value,
        },
        "3.1(c) Nil/Exempt": {
            "formula": "8_Total",
            "components": [
                {"table": "8",               "value": nil_exempt_value},
            ],
            "total": nil_exempt_value,
        },
    }

    return {"values": values, "breakdown": breakdown}


def extract_gstr3b_summary(gstr3b_data: dict) -> dict:
    sf = safe_float
    return {
        # 3.1(a) combined total
        "outward_taxable_value": sf(gstr3b_data.get("outward_taxable_value", 0)),
        "outward_igst":          sf(gstr3b_data.get("outward_igst",          0)),
        "outward_cgst":          sf(gstr3b_data.get("outward_cgst",          0)),
        "outward_sgst":          sf(gstr3b_data.get("outward_sgst",          0)),
        "outward_cess":          sf(gstr3b_data.get("outward_cess",          0)),
        # 3.2 data is nested under 'interstate_supplies' key in parse_gstr3b output
        "b2cs_value":  sf((gstr3b_data.get("interstate_supplies") or {}).get("interstate_unreg_value", 0))
                       or sf(gstr3b_data.get("interstate_unreg_value", 0)),
        "b2cs_igst":   sf((gstr3b_data.get("interstate_supplies") or {}).get("interstate_unreg_igst",  0))
                       or sf(gstr3b_data.get("interstate_unreg_igst", 0)),
        # 9B CDNR — GSTR-3B has no counterpart → always 0
        "cdnr_value":            0.0,
        "cdnr_igst":             0.0,
        # 3.1(b) exports
        "zero_rated_value":      sf(gstr3b_data.get("zero_rated_value", 0)),
        "zero_rated_igst":       sf(gstr3b_data.get("zero_rated_igst",  0)),
        # 3.1(c)
        "nil_exempt_value":      sf(gstr3b_data.get("nil_exempt_value", 0)),
        # 3.1(d) reverse charge
        "reverse_charge_value":  sf(gstr3b_data.get("inward_reverse_charge_value", 0)),
        "reverse_charge_igst":   sf(gstr3b_data.get("inward_reverse_charge_igst",  0)),
        "reverse_charge_cgst":   sf(gstr3b_data.get("inward_reverse_charge_cgst",  0)),
        "reverse_charge_sgst":   sf(gstr3b_data.get("inward_reverse_charge_sgst",  0)),
    }
# ─── Core reconciliation ──────────────────────────────────────────────────────

def reconcile_gstr1_vs_gstr3b(
    gstr1_data: dict,
    gstr3b_data: dict,
    period: str = None,
) -> Dict[str, Any]:
    """
    Compare GSTR-1 and GSTR-3B for one period.
    Returns a dict with rows, overall_status, total_variance, gstin and period.
    """

    g1_result = calculate_gstr1_liability(gstr1_data)
    g1 = g1_result["values"]
    g1_breakdown = g1_result["breakdown"]
    g3b = extract_gstr3b_summary(gstr3b_data)

    rows = []
    total_variance = 0.0

    # Keys used only for display (not compared with GSTR-3B)
    DISPLAY_ONLY_KEYS = INFO_ONLY_KEYS

    for key, section, description in RECON_ROWS:

        g1_val = g1.get(key, 0.0)
        g3b_val = g3b.get(key, 0.0)

        diff = round(g3b_val - g1_val, 2)

        if key in INFO_ONLY_KEYS:
            status = "Info"
            diff   = 0.0
        else:
            status = "Match" if abs(diff) <= AMOUNT_TOLERANCE else "Mismatch"
            total_variance += abs(diff)

        rows.append({
            "section": section,
            "description": description,
            "gstr1": g1_val,
            "gstr3b": g3b_val,
            "difference": diff,
            "status": status,
        })

    summary_comparison = {
        r["description"]: r
        for r in rows
    }

    return {
        "period": period or datetime.datetime.now().strftime("%b-%Y"),
        "gstin": gstr1_data.get("GSTIN") or gstr3b_data.get("gstin", ""),
        "rows": rows,
        "summary_comparison": summary_comparison,
        "overall_status": "Matched" if total_variance <= 10 else "Mismatch",
        "total_variance": round(total_variance, 2),
        "breakdown": g1_breakdown,
    }
# ─── Excel styles ─────────────────────────────────────────────────────────────

_GREEN  = PatternFill("solid", fgColor="D6F5D6")
_RED    = PatternFill("solid", fgColor="FFD6D6")
_INFO   = PatternFill("solid", fgColor="F0F0F0")
_HEADER = PatternFill("solid", fgColor="2D3748")
_SUB    = PatternFill("solid", fgColor="F0F4F8")
_TOTAL  = PatternFill("solid", fgColor="E2E8F0")

def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr_font(white=True):
    return Font(bold=True, color="FFFFFF" if white else "1A1A1A", size=10)

def _num_fmt(cell):
    cell.number_format = '#,##0.00'
    cell.alignment     = Alignment(horizontal="right")

def _set_col_width(ws, col_idx, width):
    ws.column_dimensions[get_column_letter(col_idx)].width = width


# ─── Monthly sheet writer ─────────────────────────────────────────────────────

def _write_period_sheet(ws, result: dict):
    """
    Write one period's full row-by-row reconciliation to a worksheet.
    Columns: Section | Description | GSTR-1 | GSTR-3B | Difference | Status
    """
    ws.freeze_panes = "A4"

    # Title
    ws.append([f"GSTR-1 vs GSTR-3B Reconciliation  —  {result['period']}"])
    ws["A1"].font = Font(bold=True, size=12, color="1A1A1A")
    if result.get("gstin"):
        ws.append([f"GSTIN: {result['gstin']}"])
        ws["A2"].font = Font(italic=True, size=10, color="555555")
    else:
        ws.append([])
    ws.append([])  # blank spacer

    # Header row
    headers = ["Section", "Description", "Formula / Tables", "GSTR-1 (₹)", "GSTR-3B (₹)", "Difference (₹)", "Status"]
    ws.append(headers)
    hrow = ws.max_row
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(hrow, ci)
        cell.fill      = _HEADER
        cell.font      = _hdr_font()
        cell.border    = _border()
        cell.alignment = Alignment(horizontal="center" if ci in (1, 6) else "right" if ci >= 3 else "left")

    # Data rows
    # Build formula lookup from breakdown
    formula_lookup: dict = {}
    for section_name, bd in result.get("breakdown", {}).items():
        tables = ", ".join(c["table"] for c in bd.get("components", []))
        formula_lookup[section_name] = f"{bd.get('formula','')}"
        
    prev_section = None
    for row in result.get("rows", []):
        section_changed = row["section"] != prev_section
        prev_section = row["section"]

        # Match breakdown key by section ref
        DESC_TO_BD = {
            "Outward Taxable Value":        "3.1(a) Total",
            "IGST":                         "3.1(a) Total",
            "CGST":                         "3.1(a) Total",
            "SGST":                         "3.1(a) Total",
            "Cess":                         "3.1(a) Total",
            "B2CS Value (Inter-state)":     "3.1(a) B2CS",
            "B2CS IGST":                    "3.1(a) B2CS",
            "CDNR Adj. Value (info)":       "3.1(a) Registered",
            "CDNR Adj. IGST (info)":        "3.1(a) Registered",
            "Exports / Zero Rated Value":   "3.1(b) Exports",
            "Exports IGST":                 "3.1(b) Exports",
            "Nil / Exempt Supplies":        "3.1(c) Nil/Exempt",
            "RCM Inward Value (3B only)":   None,
            "RCM Inward IGST (3B only)":    None,
            "RCM Inward CGST (3B only)":    None,
            "RCM Inward SGST (3B only)":    None,
        }
        bd_key      = DESC_TO_BD.get(row["description"])
        formula_str = result.get("breakdown", {}).get(bd_key, {}).get("formula", "") if bd_key else ""
        ws.append([
            row["section"],
            row["description"],
            formula_str,
            row["gstr1"],
            row["gstr3b"],
            row["difference"],
            row["status"],
        ])
        r = ws.max_row
        if row["status"] == "Match":
            fill = _GREEN
        elif row["status"] == "Info":
            fill = _INFO
        else:
            fill = _RED

        for ci in range(1, 8):
            cell = ws.cell(r, ci)
            cell.fill   = fill
            cell.border = _border()
            cell.alignment = Alignment(horizontal="right" if ci >= 4 else "left")
            if ci >= 4 and ci <= 6:
                _num_fmt(cell)
            if ci == 1 and section_changed:
                cell.font = Font(bold=True, size=9)

    # Total variance footer
    ws.append([])
    ws.append(["", "Total Variance", "", "", "", result.get("total_variance", 0),
               result.get("overall_status", "")])
    r = ws.max_row
    footer_fill = _GREEN if result.get("overall_status") == "Matched" else _RED
    for ci in range(1, 8):
        cell = ws.cell(r, ci)
        cell.fill   = footer_fill
        cell.font   = Font(bold=True)
        cell.border = _border()
        if ci == 6:
            _num_fmt(cell)

    # Column widths
    for ci, w in enumerate([10, 32, 28, 18, 18, 18, 12], 1):
        _set_col_width(ws, ci, w)


# ─── Annual summary sheet writer ──────────────────────────────────────────────

def _write_annual_sheet(ws, periods: list):
    """
    Write an annual summary sheet where every row from RECON_ROWS
    becomes a column, giving a full picture across all months.

    Layout:
      Row 1: Title
      Row 2: blank
      Row 3: Column headers — Period | GSTIN | [all RECON_ROW descriptions x3 (G1/3B/Diff)] | Total Variance | Status
      Row 4+: One row per period
      Last row: Annual totals
    """
    ws.freeze_panes = "A4"

    # Title
    ws.append(["GSTR-1 vs GSTR-3B  —  Annual Summary"])
    ws["A1"].font = Font(bold=True, size=13, color="1A1A1A")
    ws.append([])

    # Build header groups:
    # Fixed cols: Period, GSTIN
    # For each RECON_ROW: GSTR-1, GSTR-3B, Diff  (3 sub-columns)
    # Final: Total Variance, Status

    fixed_headers   = ["Period", "GSTIN"]
    row_desc_cols   = []   # (description, col_type) where col_type in G1/3B/Diff
    for _, _, desc in RECON_ROWS:
        row_desc_cols.append((desc, "GSTR-1"))
        row_desc_cols.append((desc, "GSTR-3B"))
        row_desc_cols.append((desc, "Diff"))
    tail_headers = ["Total Variance (₹)", "Status"]

    all_headers = fixed_headers + [f"{desc} ({t})" for desc, t in row_desc_cols] + tail_headers

    ws.append(all_headers)
    hrow = ws.max_row
    n_cols = len(all_headers)

    for ci in range(1, n_cols + 1):
        cell = ws.cell(hrow, ci)
        cell.fill      = _HEADER
        cell.font      = _hdr_font()
        cell.border    = _border()
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    ws.row_dimensions[hrow].height = 32  # taller header for wrapped text

    # Data rows
    annual_totals = {key: {"gstr1": 0.0, "gstr3b": 0.0, "difference": 0.0}
                     for key, _, _ in RECON_ROWS}
    annual_variance = 0.0

    for p in periods:
        by_desc = {r["description"]: r for r in p.get("rows", [])}
        row_data = [p.get("period", ""), p.get("gstin", "")]

        for _, _, desc in RECON_ROWS:
            r = by_desc.get(desc, {})
            g1_val  = r.get("gstr1", 0.0)
            g3b_val = r.get("gstr3b", 0.0)
            diff    = r.get("difference", 0.0)
            row_data += [g1_val, g3b_val, diff]
            annual_totals[next(k for k, _, d in RECON_ROWS if d == desc)]["gstr1"]      += g1_val
            annual_totals[next(k for k, _, d in RECON_ROWS if d == desc)]["gstr3b"]     += g3b_val
            annual_totals[next(k for k, _, d in RECON_ROWS if d == desc)]["difference"] += diff

        variance = p.get("total_variance", 0.0)
        annual_variance += variance
        row_data += [variance, p.get("overall_status", "")]

        ws.append(row_data)
        r_idx = ws.max_row
        is_clean = p.get("overall_status") == "Matched"
        row_fill = _GREEN if is_clean else _RED

        for ci in range(1, n_cols + 1):
            cell = ws.cell(r_idx, ci)
            cell.fill   = row_fill
            cell.border = _border()
            # Number format for numeric columns (cols 3 onwards, except last col = status)
            if ci >= 3 and ci < n_cols:
                _num_fmt(cell)

    # Annual totals footer
    totals_row = ["Annual Total", ""]
    for key, _, _ in RECON_ROWS:
        totals_row += [
            round(annual_totals[key]["gstr1"], 2),
            round(annual_totals[key]["gstr3b"], 2),
            round(annual_totals[key]["difference"], 2),
        ]
    all_matched = all(p.get("overall_status") == "Matched" for p in periods)
    totals_row += [round(annual_variance, 2), "Matched" if all_matched else "Mismatch"]

    ws.append(totals_row)
    t_idx = ws.max_row
    t_fill = _GREEN if all_matched else _RED
    for ci in range(1, n_cols + 1):
        cell = ws.cell(t_idx, ci)
        cell.fill   = t_fill
        cell.font   = Font(bold=True)
        cell.border = _border()
        if ci >= 3 and ci < n_cols:
            _num_fmt(cell)

    # Column widths
    _set_col_width(ws, 1, 14)   # Period
    _set_col_width(ws, 2, 22)   # GSTIN
    # Each RECON_ROW group = 3 cols (G1, 3B, Diff) × 16
    for i in range(len(RECON_ROWS) * 3):
        _set_col_width(ws, 3 + i, 16)
    _set_col_width(ws, 3 + len(RECON_ROWS) * 3,     16)  # Total Variance
    _set_col_width(ws, 3 + len(RECON_ROWS) * 3 + 1, 12)  # Status


# ─── Public API ───────────────────────────────────────────────────────────────
def _write_calc_summary_sheet(ws, periods: list):
    """
    Calculation Summary sheet — shows formula + components for every period.
    One block per period, then per-section breakdown.
    """
    ws.append(["GSTR-1 Calculation Summary — Formula Breakdown"])
    ws["A1"].font = Font(bold=True, size=12)
    ws.append([])

    headers = ["Period", "Section", "Formula", "Table", "Component Value (₹)",
               "Calc GSTR-1 Total (₹)", "GSTR-3B Value (₹)", "Difference (₹)", "Status"]
    ws.append(headers)
    hrow = ws.max_row
    for ci, _ in enumerate(headers, 1):
        cell = ws.cell(hrow, ci)
        cell.fill = _HEADER
        cell.font = _hdr_font()
        cell.border = _border()
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[hrow].height = 28

    for p in periods:
        breakdown = p.get("breakdown", {})
        by_desc = {r["description"]: r for r in p.get("rows", [])}
        period_label = p.get("period", "")

        for section_name, bd in breakdown.items():
            formula   = bd.get("formula", "")
            total     = bd.get("total", 0.0)
            components = bd.get("components", [])

            # Only the Total section has a direct GSTR-3B counterpart
            # Sub-sections (Registered, B2CL, B2CS) are internal breakdowns only
            SECTION_TO_DESC = {
                "3.1(a) Total":   "Outward Taxable Value",
                "3.1(a) B2CS":    "B2CS Value (Inter-state)",
                "3.1(b) Exports": "Exports / Zero Rated Value",
                "3.1(c) Nil/Exempt": "Nil / Exempt Supplies",
            }
            by_desc     = {r["description"]: r for r in p.get("rows", [])}
            desc_key    = SECTION_TO_DESC.get(section_name)
            matched_row = by_desc.get(desc_key, {}) if desc_key else {}
            g3b_val     = matched_row.get("gstr3b", None)   # None = no counterpart
            g1_val_bd   = bd.get("total", 0.0)
            if g3b_val is None:
                diff   = None
                status = "N/A"
            else:
                diff   = round(g1_val_bd - g3b_val, 2)
                status = "Match" if abs(diff) <= AMOUNT_TOLERANCE else "Mismatch"

            for i, comp in enumerate(components):
                ws.append([
                    period_label if i == 0 else "",
                    section_name if i == 0 else "",
                    formula      if i == 0 else "",
                    comp["table"],
                    comp["value"],
                    total        if i == 0 else "",
                    g3b_val      if i == 0 else "",
                    (diff   if diff   is not None else "N/A") if i == 0 else "",
                    (status if status is not None else "N/A") if i == 0 else "",
                ])
                r = ws.max_row
                fill = _GREEN if status == "Match" else (_INFO if status == "N/A" else _RED)
                for ci in range(1, 10):
                    cell = ws.cell(r, ci)
                    cell.border = _border()
                    if i == 0:
                        cell.fill = fill
                    if ci in (5, 6, 7, 8):
                        _num_fmt(cell)

        ws.append([])  # blank row between periods

    for ci, w in enumerate([14, 22, 36, 18, 18, 18, 18, 18, 12], 1):
        _set_col_width(ws, ci, w)
def generate_excel_report(periods: list, output_path: str) -> str:
    """
    Generate Excel workbook.
    - If multiple periods: Sheet 1 = Annual Summary (all RECON_ROW fields as columns),
      then one sheet per period with full row detail.
    - If single period: just one sheet with full row detail.

    `periods` is a list of reconcile_gstr1_vs_gstr3b() result dicts.
    """
    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    if len(periods) > 1:
        ws_annual = wb.create_sheet("Annual Summary")
        _write_annual_sheet(ws_annual, periods)

    ws_calc = wb.create_sheet("Calculation Summary")
    _write_calc_summary_sheet(ws_calc, periods)

    for p in periods:
        sheet_name = (p.get("period") or "Period")[:31]
        ws = wb.create_sheet(sheet_name)
        _write_period_sheet(ws, p)

    wb.save(output_path)
    logger.info(f"Excel report saved: {output_path}")
    return output_path