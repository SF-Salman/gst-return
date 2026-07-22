import re
import logging
from typing import Optional
import pandas as pd

logger = logging.getLogger(__name__)

# ─── Required template headers (exact case-insensitive match, trimmed) ───────

REQUIRED_HEADERS = {
    "month":         "month",
    "category":      "category",
    "taxable value": "taxable_value",
    "igst":          "igst",
    "cgst":          "cgst",
    "sgst":          "sgst",
    "cess":          "cess",
}

OPTIONAL_HEADERS = {
    "voucher date": "voucher_date",
    "voucher no":   "voucher_no",
    "ledger":       "ledger",
    "type":         "type",
    "remarks":      "remarks",
}

ALL_HEADERS = {**REQUIRED_HEADERS, **OPTIONAL_HEADERS}

VALID_CATEGORIES = {"OUTPUT", "ITC", "RCM", "REVERSAL", "EXEMPT"}

# ─── Helpers ──────────────────────────────────────────────────────────────────

def _norm_header(s: str) -> str:
    """Case-insensitive, whitespace-trimmed header normalization."""
    return re.sub(r"\s+", " ", str(s).strip().lower())

def _safe_float(v) -> float:
    try:
        if pd.isna(v):
            return 0.0
        return round(float(str(v).replace(",", "").strip()), 2)
    except Exception:
        return 0.0

def _parse_period(val) -> Optional[str]:
    """Convert a cell value to YYYY-MM string. Accepts Apr-25, April 2025, 04/2025, 2025-04, etc."""
    if pd.isna(val):
        return None
    s = str(val).strip()

    if re.match(r"^\d{4}-\d{2}$", s):
        return s

    MONTH_NUM = {
        "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
        "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
    }

    # "Apr-25", "Apr-2025", "April 2025", "April-25"
    m = re.match(r"^([A-Za-z]{3,9})[-\s](\d{2,4})$", s)
    if m:
        mon_key = m.group(1).lower()[:3]
        if mon_key in MONTH_NUM:
            year = m.group(2)
            year = "20" + year if len(year) == 2 else year
            return f"{year}-{MONTH_NUM[mon_key]:02d}"

    # "04/2025", "04-2025"
    m2 = re.match(r"^(\d{1,2})[/-](\d{4})$", s)
    if m2:
        return f"{m2.group(2)}-{int(m2.group(1)):02d}"

    # Last resort: pandas, but only if the above patterns didn't match
    try:
        dt = pd.to_datetime(s, dayfirst=True)
        return dt.strftime("%Y-%m")
    except Exception:
        pass

    return None

def _find_template_sheet(df_map: dict[str, "pd.DataFrame"]) -> Optional[tuple[str, "pd.DataFrame", dict[str, str]]]:
    """
    Scan all sheets for one whose header row matches all REQUIRED_HEADERS
    (case-insensitive, trimmed). Returns (sheet_name, df, column_map) or None.
    column_map maps normalized field name -> actual column name in df.
    """
    for sheet_name, df in df_map.items():
        if df.empty or len(df.columns) < len(REQUIRED_HEADERS):
            continue

        header_lookup = {_norm_header(c): c for c in df.columns}

        column_map: dict[str, str] = {}
        missing = []
        for header_key, field_name in ALL_HEADERS.items():
            actual_col = header_lookup.get(header_key)
            if actual_col is not None:
                column_map[field_name] = actual_col
            elif header_key in REQUIRED_HEADERS:
                missing.append(header_key)

        if not missing:
            return sheet_name, df, column_map

    return None

# ─── NEW section-based template (Month-per-row) ───────────────────────────────
#
# Layout (as produced by generate_books_template()): two header rows, then one
# data row per month (Apr-25 ... Mar-26), then a "Total" row (ignored on read —
# totals are always recalculated from the underlying figures).

# section_key -> (keywords to match in the row-1 section header, ordered field keys)
SECTION_RULES: list[tuple[str, list[str], list[str]]] = [
    ("outward",         ["outward supply", "tax on outward"],                ["taxable_value", "zero_rated", "nil_rated", "igst", "cgst", "sgst"]),
    ("rcm",              ["rcm", "reverse charge", "inward supply"],          ["taxable_value", "igst", "cgst", "sgst"]),
    ("itc_reversed",     ["itc reversed", "reversed"],                        ["igst", "cgst", "sgst"]),
    ("ineligible_itc",   ["ineligible"],                                      ["igst", "cgst", "sgst"]),
    ("net_itc",          ["net itc"],                                        ["igst", "cgst", "sgst"]),
    ("itc_available",    ["itc available"],                                  ["igst", "cgst", "sgst"]),
    ("itc_utilised",     ["itc utilised", "itc utilized"],                   ["igst", "cgst", "sgst"]),
    ("net_tax_payable",  ["net tax payable"],                                 ["igst", "cgst", "sgst"]),
    ("late_fee",         ["late fee"],                                       ["cgst", "sgst"]),
    ("cash_paid",        ["paid in cash", "paid through cash", "cash"],       ["igst", "cgst", "sgst"]),
    ("difference",       ["difference", "short", "excess"],                  ["igst", "cgst", "sgst"]),
]
# Columns whose row-1 header is just a running total (recalculated, never read)
_SKIP_SECTION_HINTS = ["total (a)", "total (b)", "total (c)", "total tax payable", "total itc utilised"]

_FIELD_HINTS = {
    "taxable_value": ["taxable value"],
    "zero_rated":    ["zero rated"],
    "nil_rated":     ["nil rated"],
    "igst":          ["igst"],
    "cgst":          ["cgst"],
    "sgst":          ["sgst"],
}

_MONTH_ROW_RE = re.compile(r"^[A-Za-z]{3,9}[-\s]\d{2,4}$")


def _classify_column(section_text: str, field_text: str) -> Optional[tuple[str, str]]:
    """Given normalized (section-header, field-header) text for a column,
    return (section_key, field_key) or None if the column should be ignored
    (e.g. a running total, or an unrecognized column)."""
    if any(hint in section_text for hint in _SKIP_SECTION_HINTS):
        return None
    if section_text.strip() == "total" or field_text.strip() == "total":
        return None

    matched_section = None
    for key, hints, _fields in SECTION_RULES:
        if any(hint in section_text for hint in hints):
            matched_section = key
            break
    if not matched_section:
        return None

    for field_key, hints in _FIELD_HINTS.items():
        if any(hint in field_text for hint in hints):
            if field_key in ("taxable_value", "zero_rated", "nil_rated") and matched_section not in ("outward", "rcm"):
                continue
            return matched_section, field_key
    return None


def _find_section_sheet(raw_map: dict) -> Optional[tuple[str, "pd.DataFrame"]]:
    """Detect a sheet using the new two-header-row, Month-per-row layout.
    Returns (sheet_name, raw_df) with raw_df read with header=None, or None."""
    for sheet_name, raw in raw_map.items():
        if raw.shape[0] < 4 or raw.shape[1] < 8:
            continue

        row1 = raw.iloc[0].ffill()
        row2 = raw.iloc[1]

        first_cell = _norm_header(row2.iloc[0]) if pd.notna(row2.iloc[0]) else _norm_header(row1.iloc[0])
        if "month" not in first_cell:
            continue

        found_month_row = False
        for r in range(2, min(raw.shape[0], 20)):
            v = raw.iat[r, 0]
            if pd.notna(v) and _MONTH_ROW_RE.match(str(v).strip()):
                found_month_row = True
                break
        if not found_month_row:
            continue

        recognized = 0
        for c in range(1, raw.shape[1]):
            sec = _norm_header(row1.iloc[c]) if pd.notna(row1.iloc[c]) else ""
            fld = _norm_header(row2.iloc[c]) if pd.notna(row2.iloc[c]) else ""
            if _classify_column(sec, fld):
                recognized += 1
        if recognized >= 10:
            return sheet_name, raw

    return None


def _extract_books_new_format(sheet_name: str, raw: "pd.DataFrame") -> dict:
    row1 = raw.iloc[0].ffill()
    row2 = raw.iloc[1]

    col_map: dict[int, tuple[str, str]] = {}
    for c in range(1, raw.shape[1]):
        sec = _norm_header(row1.iloc[c]) if pd.notna(row1.iloc[c]) else ""
        fld = _norm_header(row2.iloc[c]) if pd.notna(row2.iloc[c]) else ""
        cls = _classify_column(sec, fld)
        if cls:
            col_map[c] = cls

    periods: dict[str, dict] = {}
    rows_processed = 0
    rows_skipped = 0

    for r in range(2, raw.shape[0]):
        month_cell = raw.iat[r, 0]
        if pd.isna(month_cell) or not _MONTH_ROW_RE.match(str(month_cell).strip()):
            rows_skipped += 1
            continue  # skip blank rows and the trailing "Total" row

        period_key = _parse_period(month_cell)
        if not period_key:
            rows_skipped += 1
            continue

        sec_vals: dict[str, dict[str, float]] = {}
        for c, (sec_key, field_key) in col_map.items():
            sec_vals.setdefault(sec_key, {})[field_key] = _safe_float(raw.iat[r, c])

        def g(sec, field):
            return sec_vals.get(sec, {}).get(field, 0.0)

        outward_igst  = g("outward", "igst")
        outward_cgst  = g("outward", "cgst")
        outward_sgst  = g("outward", "sgst")
        itc_av_igst   = g("itc_available", "igst")
        itc_av_cgst   = g("itc_available", "cgst")
        itc_av_sgst   = g("itc_available", "sgst")
        itc_rev_igst  = g("itc_reversed", "igst")
        itc_rev_cgst  = g("itc_reversed", "cgst")
        itc_rev_sgst  = g("itc_reversed", "sgst")

        periods[period_key] = {
            "taxable_value":      g("outward", "taxable_value"),
            "zero_rated_value":   g("outward", "zero_rated"),
            "nil_rated_value":    g("outward", "nil_rated"),
            "igst_output":        outward_igst,
            "cgst_output":        outward_cgst,
            "sgst_output":        outward_sgst,
            "rcm_taxable":        g("rcm", "taxable_value"),
            "rcm_igst":           g("rcm", "igst"),
            "rcm_cgst":           g("rcm", "cgst"),
            "rcm_sgst":           g("rcm", "sgst"),
            "itc_reversed_igst":  itc_rev_igst,
            "itc_reversed_cgst":  itc_rev_cgst,
            "itc_reversed_sgst":  itc_rev_sgst,
            "ineligible_igst":    g("ineligible_itc", "igst"),
            "ineligible_cgst":    g("ineligible_itc", "cgst"),
            "ineligible_sgst":    g("ineligible_itc", "sgst"),
            "net_itc_igst":       g("net_itc", "igst"),
            "net_itc_cgst":       g("net_itc", "cgst"),
            "net_itc_sgst":       g("net_itc", "sgst"),
            "itc_util_igst":      g("itc_utilised", "igst"),
            "itc_util_cgst":      g("itc_utilised", "cgst"),
            "itc_util_sgst":      g("itc_utilised", "sgst"),
            "net_tax_payable":    g("net_tax_payable", "igst") + g("net_tax_payable", "cgst") + g("net_tax_payable", "sgst"),
            "cash_paid_igst":     g("cash_paid", "igst"),
            "cash_paid_cgst":     g("cash_paid", "cgst"),
            "cash_paid_sgst":     g("cash_paid", "sgst"),
            "interest_igst":      0.0,
            "interest_cgst":      0.0,
            "interest_sgst":      0.0,
            "late_fee_cgst":      g("late_fee", "cgst"),
            "late_fee_sgst":      g("late_fee", "sgst"),
            # NOTE: itc_igst/itc_cgst/itc_sgst are RAW "ITC Available" figures
            # (reversal is tracked separately via itc_reversed_igst/cgst/sgst)
            # — this matches what reconcile_period() / RECON_ROWS expect.
            "itc_igst":           itc_av_igst,
            "itc_cgst":           itc_av_cgst,
            "itc_sgst":           itc_av_sgst,
            # Aggregate fields kept for backward compatibility with GSTR-2B vs Books
            "sales":              g("outward", "taxable_value"),
            "igst":               outward_igst,
            "cgst":               outward_cgst,
            "sgst":               outward_sgst,
            "cess":               0.0,
            "itc":                round((itc_av_igst + itc_av_cgst + itc_av_sgst) - (itc_rev_igst + itc_rev_cgst + itc_rev_sgst), 2),
            "rcm":                round(g("rcm", "igst") + g("rcm", "cgst") + g("rcm", "sgst"), 2),
        }
        rows_processed += 1

    for p in periods:
        periods[p] = {k: round(v, 2) for k, v in periods[p].items()}

    audit = {
        "sheet_used": sheet_name,
        "rows_processed": rows_processed,
        "rows_skipped": rows_skipped,
        "format": "section-based (new template)",
    }
    return {"periods": periods, "audit": audit}


# ─── Main extractor ───────────────────────────────────────────────────────────

def extract_books(file_bytes: bytes, filename: str) -> dict:
    """
    Extract Books data. Two supported template formats are auto-detected
    (no filename/sheet-name assumptions — only header content is inspected):

    1. NEW section-based format (Month-per-row, two header rows with grouped
       sections like "Tax on Outward Supply", "ITC Available", etc.) — this is
       the format produced by generate_books_template().
    2. LEGACY row-per-voucher format (Month, Category, Taxable Value, IGST,
       CGST, SGST, CESS columns) — kept for backward compatibility with any
       older Books files already in use (e.g. GSTR-2B vs Books uploads).

    Returns the same "periods" / "audit" shape either way, so callers
    (reconcile_gstr3b_vs_books, reconcile_gstr2b_vs_books) do not need to change.
    """
    fname = filename.lower()

    if fname.endswith(".csv"):
        raw_map = {"Sheet1": pd.read_csv(pd.io.common.BytesIO(file_bytes), dtype=str, header=None)}
    else:
        xl = pd.ExcelFile(pd.io.common.BytesIO(file_bytes))
        raw_map = {sheet: xl.parse(sheet, dtype=str, header=None) for sheet in xl.sheet_names}

    new_format = _find_section_sheet(raw_map)
    if new_format:
        return _extract_books_new_format(*new_format)

    return _extract_books_legacy(file_bytes, filename)


def _extract_books_legacy(file_bytes: bytes, filename: str) -> dict:
    """
    Extract Books data using the strict legacy template format.
    Scans all sheets in the workbook for one matching the required headers
    (case-insensitive, whitespace-trimmed). The sheet/file name itself does
    not matter — only the header row content.

    Returns:
    {
      "periods": { "YYYY-MM": { "sales", "taxable_value", "igst", "cgst", "sgst", "cess", "itc", "rcm" } },
      "audit": { "sheet_used": str, "rows_processed": int, "rows_skipped": int, "categories_seen": {...} }
    }

    Raises ValueError if no sheet matches the required template headers.
    """
    fname = filename.lower()

    if fname.endswith(".csv"):
        df_map = {"Sheet1": pd.read_csv(pd.io.common.BytesIO(file_bytes), dtype=str)}
    else:
        xl = pd.ExcelFile(pd.io.common.BytesIO(file_bytes))
        df_map = {sheet: xl.parse(sheet, dtype=str) for sheet in xl.sheet_names}

    found = _find_template_sheet(df_map)
    if not found:
        required_list = ", ".join(sorted(REQUIRED_HEADERS.keys()))
        raise ValueError(
            f"Could not find required columns ({required_list}) in any sheet. "
            f"Please use the downloaded Books template format."
        )

    sheet_name, df, column_map = found

    periods: dict[str, dict] = {}
    categories_seen: dict[str, int] = {}
    rows_processed = 0
    rows_skipped = 0

    def _ensure_period(p: str):
        if p not in periods:
            periods[p] = {
                "sales": 0.0, "taxable_value": 0.0,
                "igst": 0.0, "cgst": 0.0, "sgst": 0.0,
                "cess": 0.0, "itc": 0.0, "rcm": 0.0,
                # Additive, tax-head-split ITC figures — used by GSTR-2B vs Books.
                # Existing 'itc' (combined) field above is untouched for backward
                # compatibility with GSTR-3B vs Books.
                "itc_igst": 0.0, "itc_cgst": 0.0, "itc_sgst": 0.0,
            }

    col_month    = column_map["month"]
    col_category = column_map["category"]
    col_taxable  = column_map["taxable_value"]
    col_igst     = column_map["igst"]
    col_cgst     = column_map["cgst"]
    col_sgst     = column_map["sgst"]
    col_cess     = column_map["cess"]

    for _, row in df.iterrows():
        period_key = _parse_period(row[col_month])
        category_raw = str(row[col_category]).strip().upper() if pd.notna(row[col_category]) else ""

        if not period_key or not category_raw:
            rows_skipped += 1
            continue

        if category_raw not in VALID_CATEGORIES:
            rows_skipped += 1
            continue

        _ensure_period(period_key)
        categories_seen[category_raw] = categories_seen.get(category_raw, 0) + 1

        taxable = _safe_float(row[col_taxable])
        igst    = _safe_float(row[col_igst])
        cgst    = _safe_float(row[col_cgst])
        sgst    = _safe_float(row[col_sgst])
        cess    = _safe_float(row[col_cess])

        if category_raw == "OUTPUT":
            periods[period_key]["sales"] += taxable
            periods[period_key]["taxable_value"] += taxable
            periods[period_key]["igst"] += igst
            periods[period_key]["cgst"] += cgst
            periods[period_key]["sgst"] += sgst
            periods[period_key]["cess"] += cess

        elif category_raw == "ITC":
            periods[period_key]["itc"] += igst + cgst + sgst + cess
            periods[period_key]["itc_igst"] += igst
            periods[period_key]["itc_cgst"] += cgst
            periods[period_key]["itc_sgst"] += sgst

        elif category_raw == "REVERSAL":
            periods[period_key]["itc"] -= (igst + cgst + sgst + cess)
            periods[period_key]["itc_igst"] -= igst
            periods[period_key]["itc_cgst"] -= cgst
            periods[period_key]["itc_sgst"] -= sgst

        elif category_raw == "RCM":
            periods[period_key]["rcm"] += igst + cgst + sgst + cess

        elif category_raw == "EXEMPT":
            # Tracked but not added to taxable/tax totals — informational only
            pass

        rows_processed += 1

    for p in periods:
        periods[p] = {k: round(v, 2) for k, v in periods[p].items()}

    audit = {
        "sheet_used": sheet_name,
        "rows_processed": rows_processed,
        "rows_skipped": rows_skipped,
        "categories_seen": categories_seen,
    }

    return {"periods": periods, "audit": audit}


# ─── Template generator ───────────────────────────────────────────────────────

_TEMPLATE_SECTIONS = [
    ("Tax on Outward Supply",              ["Taxable Value", "Zero Rated", "Nil Rated", "IGST", "CGST", "SGST"]),
    ("Total (A)",                          ["Total (A)"]),
    ("Outward Taxable Supply liable to RCM", ["Taxable Value", "IGST", "CGST", "SGST"]),
    ("Total (B)",                          ["Total (B)"]),
    ("ITC Available",                      ["IGST", "CGST", "SGST"]),
    ("ITC Reversed",                       ["IGST", "CGST", "SGST"]),
    ("Ineligible ITC",                     ["IGST", "CGST", "SGST"]),
    ("Net ITC Available",                  ["IGST", "CGST", "SGST"]),
    ("Total (C)",                          ["Total (C)"]),
    ("ITC Utilised",                       ["IGST", "CGST", "SGST"]),
    ("Net Tax Payable",                    ["IGST", "CGST", "SGST"]),
    ("Total Tax Payable [Net]",            ["Total"]),
    ("Amount Paid in Cash",                ["IGST", "CGST", "SGST"]),
    ("Total",                              ["Total"]),
    ("Difference short/(excess)",          ["IGST", "CGST", "SGST"]),
    ("Late Fee Paid Through Cash",         ["CGST", "SGST"]),
]

_MONTHS_FY = ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]


def _build_section_sheet(ws, fill_sample: bool):
    """Write the two-row section header + 12 months of data rows (blank, or
    filled with sample figures) + a Total row with SUM formulas. Shared by
    the template and sample generators so both stay in sync."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="2D3748")
    sub_fill    = PatternFill("solid", fgColor="4A5568")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.cell(1, 1, "Month")
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    col = 2
    for section_label, fields in _TEMPLATE_SECTIONS:
        start_col = col
        for field in fields:
            ws.cell(2, col, field)
            col += 1
        end_col = col - 1
        ws.cell(1, start_col, section_label)
        if end_col > start_col:
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

    total_cols = col - 1
    for r in (1, 2):
        for ci in range(1, total_cols + 1):
            c = ws.cell(r, ci)
            c.font = Font(bold=True, color="FFFFFF", size=9)
            c.fill = header_fill if r == 1 else sub_fill
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = border

    ws.freeze_panes = "B3"
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 24
    ws.column_dimensions["A"].width = 10
    for ci in range(2, total_cols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 12

    sample_month = {
        "Tax on Outward Supply": {"Taxable Value": 6000000, "Zero Rated": 0, "Nil Rated": 0, "IGST": 0, "CGST": 54000, "SGST": 54000},
        "Outward Taxable Supply liable to RCM": {"Taxable Value": 60000, "IGST": 0, "CGST": 5400, "SGST": 5400},
        "ITC Available": {"IGST": 0, "CGST": 15000, "SGST": 15000},
        "ITC Reversed": {"IGST": 0, "CGST": 0, "SGST": 0},
        "Ineligible ITC": {"IGST": 0, "CGST": 0, "SGST": 0},
        "Net ITC Available": {"IGST": 0, "CGST": 15000, "SGST": 15000},
        "ITC Utilised": {"IGST": 0, "CGST": 15000, "SGST": 15000},
        "Net Tax Payable": {"IGST": 0, "CGST": 44400, "SGST": 44400},
        "Amount Paid in Cash": {"IGST": 0, "CGST": 44400, "SGST": 44400},
        "Difference short/(excess)": {"IGST": 0, "CGST": 0, "SGST": 0},
        "Late Fee Paid Through Cash": {"CGST": 0, "SGST": 0},
    }

    data_start_row = 3
    for mi, mon in enumerate(_MONTHS_FY):
        yy = 25 if mon in ("Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec") else 26
        r = data_start_row + mi
        ws.cell(r, 1, f"{mon}-{yy}")
        col = 2
        for section_label, fields in _TEMPLATE_SECTIONS:
            for field in fields:
                if section_label.startswith("Total"):
                    col += 1
                    continue
                val = sample_month.get(section_label, {}).get(field, 0) if fill_sample else 0
                ws.cell(r, col, val)
                ws.cell(r, col).number_format = '#,##0'
                col += 1
        for ci in range(1, total_cols + 1):
            ws.cell(r, ci).border = border

    last_data_row = data_start_row + len(_MONTHS_FY) - 1
    total_row = last_data_row + 1
    ws.cell(total_row, 1, "Total")
    ws.cell(total_row, 1).font = Font(bold=True)

    col = 2
    for section_label, fields in _TEMPLATE_SECTIONS:
        for field in fields:
            letter = get_column_letter(col)
            if section_label.startswith("Total"):
                col += 1
                continue
            formula = f"=SUM({letter}{data_start_row}:{letter}{last_data_row})"
            ws.cell(total_row, col, formula)
            ws.cell(total_row, col).number_format = '#,##0'
            ws.cell(total_row, col).font = Font(bold=True)
            col += 1

    for ci in range(1, total_cols + 1):
        ws.cell(total_row, ci).border = border

    return total_cols


def generate_books_template() -> bytes:
    """Generate the downloadable BOOKS_DATA template (blank, new section-based
    Month-per-row layout matching the GSTR-3B vs Books screen)."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "BOOKS_DATA"
    _build_section_sheet(ws, fill_sample=False)

    readme = wb.create_sheet("README")
    readme.append(["Books Template — Instructions"])
    readme["A1"].font = Font(bold=True, size=13)
    readme.append([])
    readme.append(["Fill in one row per month (Apr-25 ... Mar-26)."])
    readme.append(["Do not edit the Month column format (e.g. Apr-25, May-25...)."])
    readme.append(["Leave the Total row as-is — its formulas recalculate automatically."])
    readme.append([])
    readme.append(["Sections, left to right:"])
    readme.cell(readme.max_row, 1).font = Font(bold=True)
    for section_label, fields in _TEMPLATE_SECTIONS:
        readme.append(["", section_label, ", ".join(fields)])
    readme.column_dimensions["A"].width = 4
    readme.column_dimensions["B"].width = 38
    readme.column_dimensions["C"].width = 50

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def generate_books_sample() -> bytes:
    """Generate a filled-in sample BOOKS_DATA workbook (same layout as the
    template) so users can see the expected format with realistic figures."""
    import io
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "BOOKS_DATA"
    _build_section_sheet(ws, fill_sample=True)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()