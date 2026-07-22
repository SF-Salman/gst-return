from fastapi import FastAPI, UploadFile, File, Query, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse, FileResponse, Response
from fastapi.staticfiles import StaticFiles
from backend.core.gstr3b_books import reconcile_gstr3b_vs_books, generate_gstr3b_books_excel
from backend.core.books_extractor import extract_books
from backend.core.gstr2b_extractor import extract_gstr2b, merge_gstr2b_files
from backend.core.gstr2b_vs_books import reconcile_gstr2b_vs_books, generate_gstr2b_books_excel
import tempfile
import os
import io
import json
import logging
import datetime
import decimal
import re
from collections import OrderedDict, defaultdict
from openpyxl import Workbook
from backend.core.gstr1_vs_3b import reconcile_gstr1_vs_gstr3b, generate_excel_report
from gstr3b import parse_gstr3b, create_excel
from gstr1 import extract_gstr1_data
from json_import import extract_gstr1_from_json, extract_gstr3b_from_json
from backend.core.reconciliation import reconcile
from backend.core.extractors.purchase_register import extract_purchase_register
from backend.core.extractors.gstr2b import extract_2b
from typing import Optional, List
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

app = FastAPI(title="GST Returns Extractor API", version="1.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5173",
        "http://127.0.0.1:5173",
        "http://localhost:5174",
        "http://127.0.0.1:5174",
        "http://localhost:5175",
        "http://127.0.0.1:5175",
        "http://localhost:5176",
        "http://127.0.0.1:5176",
        "http://localhost:5177",
        "http://127.0.0.1:5177",
        "http://localhost:3000",
        "http://127.0.0.1:3000",
        "http://localhost:8000",     
        "http://127.0.0.1:8000",  
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Lightweight health endpoint for orchestration and uptime checks
@app.get("/api/health")
def health():
    return {"status": "ok", "time": datetime.datetime.utcnow().isoformat()}
@app.post("/api/reconcile")
async def reconcile_endpoint(payload: dict):
    # Validate required keys exist before doing anything
    source = payload.get("source")
    target_raw = payload.get("target_2a")
    pr_raw = payload.get("target_pr")  # optional — purchase register records (pre-parsed)

    if not source:
        return JSONResponse(status_code=400, content={"error": "Missing 'source' records."})
    if not target_raw and not pr_raw:
        return JSONResponse(status_code=400, content={"error": "Missing target — provide 'target_2a' or 'target_pr'."})

    try:
        if target_raw:
            target = extract_2b(target_raw)
        else:
            target = pr_raw  # already normalized InvoiceRecord list from /api/purchase_register
        return reconcile(source, target)
    except KeyError as e:
        return JSONResponse(status_code=400, content={"error": f"Missing field in payload: {e}"})
    except Exception as e:
        logger.error("Reconcile failed: %s", e, exc_info=True)
        return JSONResponse(status_code=500, content={"error": "Reconciliation failed. Check server logs."})


@app.post("/api/purchase_register")
async def parse_purchase_register(file: UploadFile = File(...)):
    """
    Upload an Excel or CSV purchase register.
    Returns a list of normalized InvoiceRecord dicts.
    The frontend passes this list back in the /api/reconcile payload as 'target_pr'.
    """
    try:
        content = await file.read()
        records = extract_purchase_register(content, file.filename or "upload")
        return {"records": records, "count": len(records)}
    except ValueError as e:
        return JSONResponse(status_code=400, content={"error": str(e)})
    except Exception as e:
        logger.error("Purchase register parse failed: %s", e, exc_info=True)
        return JSONResponse(status_code=500, content={"error": "Could not parse file."})

@app.get("/favicon.ico")
def favicon():
    png_path = os.path.join(DIST_DIR, "icon-2024.png")
    if os.path.exists(png_path):
        return FileResponse(png_path, media_type="image/png")
    # Fallback to public during dev
    alt_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "frontend", "public", "icon-2024.png"))
    if os.path.exists(alt_path):
        return FileResponse(alt_path, media_type="image/png")
    return JSONResponse(status_code=404, content={"error": "favicon not found"})

# Frontend mount deferred until after API routes are registered (see bottom)
DIST_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "frontend", "dist"))


def _json_safe(value):
    if isinstance(value, (datetime.date, datetime.datetime)):
        return value.isoformat()
    if isinstance(value, decimal.Decimal):
        # Prefer float for numeric fields; fallback to string if NaN/Inf
        try:
            return float(value)
        except Exception:
            return str(value)
    if isinstance(value, bytes):
        # Represent bytes as hex string to avoid binary issues
        return value.hex()
    return value


def to_jsonable(obj):
    if isinstance(obj, dict):
        return {str(k): to_jsonable(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple, set)):
        return [to_jsonable(v) for v in obj]
    return _json_safe(obj)


@app.post("/api/gstr3b/pdf")
async def parse_gstr3b_pdf(file: UploadFile = File(...)):
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_file:
            content = await file.read()
            tmp_file.write(content)
            tmp_path = tmp_file.name

        data = parse_gstr3b(tmp_path)
        data['filename'] = file.filename
        os.unlink(tmp_path)
        return JSONResponse(content=to_jsonable(data))
    except Exception as e:
        logger.exception("Failed to parse GSTR-3B PDF")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.post("/api/gstr1/pdf")
async def parse_gstr1_pdf(file: UploadFile = File(...)):
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_file:
            content = await file.read()
            tmp_file.write(content)
            tmp_path = tmp_file.name

        data = extract_gstr1_data(tmp_path)
        data['filename'] = file.filename
        os.unlink(tmp_path)
        return JSONResponse(content=to_jsonable(data))
    except Exception as e:
        logger.exception("Failed to parse GSTR-1 PDF")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.post("/api/gstr3b/json")
async def parse_gstr3b_json(payload: dict):
    try:
        data = extract_gstr3b_from_json(payload)
        return JSONResponse(content=to_jsonable(data))
    except Exception as e:
        logger.exception("Failed to parse GSTR-3B JSON")
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.post("/api/gstr2b/json")
async def parse_gstr2b_json(payload: dict):
    """
    Parse a GSTR-2B JSON file downloaded from the GSTN portal.
    Returns a list of normalized InvoiceRecord dicts with source='2B'.
    These can be passed directly to /api/reconcile as 'target_2a' payload
    (the reconciliation engine handles both 2A and 2B records the same way).
    """
    try:
        records = extract_2b(payload)
        return JSONResponse(content=to_jsonable({
            "records": records,
            "count": len(records),
            "source": "2B"
        }))
    except Exception as e:
        logger.exception("Failed to parse GSTR-2B JSON")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.post("/api/gstr1/json")
async def parse_gstr1_json(payload: dict):
    try:
        data = extract_gstr1_from_json(payload)
        return JSONResponse(content=to_jsonable(data))
    except Exception as e:
        logger.exception("Failed to parse GSTR-1 JSON")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.post("/api/excel")
async def export_excel(records: list[dict]):
    try:
        output: io.BytesIO = create_excel(records)
        output.seek(0)
        headers = {
            "Content-Disposition": "attachment; filename=returns_data.xlsx",
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        }
        return StreamingResponse(output, headers=headers)
    except Exception as e:
        logger.exception("Failed to create Excel export")
        return JSONResponse(status_code=500, content={"error": str(e)})


# -----------------------------
# Summarization Helpers & API
# -----------------------------

def _normalize_key(s: str) -> str:
    try:
        return re.sub(r"[^a-z0-9]", "", str(s).lower()).strip()
    except Exception:
        return str(s)

# Treat common sentinel values as missing and provide normalization helpers
def _is_missing_value(s: str) -> bool:
    if s is None:
        return True
    v = str(s).strip().lower()
    return v in ("", "none", "null", "n/a", "na", "-", "—")

def _normalize_category(cat: str, heading: str) -> str:
    cat_str = (cat or "").strip()
    if _is_missing_value(cat_str):
        return (heading or "").strip()
    return cat_str

def _normalize_optional(s: str) -> str | None:
    if s is None:
        return None
    s_str = str(s).strip()
    if _is_missing_value(s_str):
        return None
    return s_str


def _get_period_label(record: dict, return_type: str) -> str:
    # Attempt to derive a user-friendly period label
    # Prefer explicit period strings; fall back to financial year where appropriate
    for k in ("TaxPeriod", "tax_period", "period", "ret_period"):
        v = record.get(k)
        if v:
            return str(v)
    if return_type == "GSTR-1":
        for k in ("FinancialYear", "financial_year", "fy", "year"):
            v = record.get(k)
            if v:
                return str(v)
    # Last resort: filename or index
    return str(record.get("filename", ""))

def _month_order_index(label: str) -> int | None:
    """Return index 0..11 for April..March if label contains a matching month, else None."""
    months = [
        ("april", ["apr", "april"]),
        ("may", ["may"]),
        ("june", ["jun", "june"]),
        ("july", ["jul", "july"]),
        ("august", ["aug", "august"]),
        ("september", ["sep", "sept", "september"]),
        ("october", ["oct", "october"]),
        ("november", ["nov", "november"]),
        ("december", ["dec", "december"]),
        ("january", ["jan", "january"]),
        ("february", ["feb", "february"]),
        ("march", ["mar", "march"]),
    ]
    lc = str(label).lower()
    for i, (_m, syns) in enumerate(months):
        for s in syns:
            if s in lc:
                return i
    return None

def _sort_summary_columns(summary: dict) -> dict:
    """Sort columns (Period/TaxPeriod) April..March and reorder row values accordingly."""
    cols = list(summary.get("columns", []))
    if not cols:
        return summary
    items = [(idx, c, _month_order_index(c)) for idx, c in enumerate(cols)]
    known = [it for it in items if it[2] is not None]
    unknown = [it for it in items if it[2] is None]
    known.sort(key=lambda x: (x[2], x[0]))  # by month index, stable by original index
    order = [idx for idx, _c, _m in known] + [idx for idx, _c, _m in unknown]
    # Apply reorder to columns
    summary["columns"] = [cols[i] for i in order]
    # Apply reorder to values in each row
    for section in summary.get("sections", []):
        for cat in section.get("categories", []):
            if cat.get("subcategories"):
                for sub in cat.get("subcategories", []):
                    for row in sub.get("rows", []):
                        vals = row.get("values", [])
                        row["values"] = [vals[i] if i < len(vals) else None for i in order]
            for row in (cat.get("rows", []) or []):
                vals = row.get("values", [])
                row["values"] = [vals[i] if i < len(vals) else None for i in order]
    return summary


def _load_schema(return_type: str) -> list[dict]:
    """
    Load schema rows from JSON files moved into backend/schemas.
    Supports keys:
      - Headings / Heading -> heading
      - Category -> category
      - Sub Category / Subcategory -> subcategory (optional for GSTR-3B)
      - Field Name / Field / Key -> field_name
      - Field Description / Description / Label -> field_description
      - Column Index / Index / Order -> column_index (optional)
    """
    file_map = {
        "GSTR-1": "gstr1.json",
        "GSTR-3B": "gstr3b.json",
    }
    filename = file_map.get(return_type)
    if not filename:
        raise FileNotFoundError(f"Unknown return type for schema: {return_type}")

    # Prefer backend/schemas, fallback to project root if not found
    base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    schema_path = os.path.join(base_dir, "schemas", filename)
    if not os.path.exists(schema_path):
        project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
        alt_path = os.path.join(project_root, filename)
        if os.path.exists(alt_path):
            schema_path = alt_path
        else:
            raise FileNotFoundError(f"Schema JSON not found: {schema_path}")

    with open(schema_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    rows: list[dict] = []
    for item in data if isinstance(data, list) else []:
        def pick(*keys):
            for k in keys:
                v = item.get(k)
                if v is not None:
                    return v
            return None

        heading = pick("Headings", "Heading")
        category = pick("Category", "Table", "Group")
        subcategory = pick("Sub Category", "Subcategory", "Subcat")
        if isinstance(subcategory, str) and _normalize_key(subcategory) in {"none", "null", ""}:
            subcategory = None
        field_name = pick("Field Name", "Field", "Key", "Column Name", "Name")
        field_description = pick("Field Description", "Description", "Label", "Desc")
        column_index_raw = pick("Column Index", "Index", "Order", "Idx", "Column")
        column_index = None
        if column_index_raw is not None and str(column_index_raw).strip() != "":
            try:
                column_index = int(column_index_raw)
            except Exception:
                try:
                    column_index = int(float(column_index_raw))
                except Exception:
                    column_index = None

        if field_name:
            rows.append({
                "heading": heading,
                "category": category,
                "subcategory": subcategory,
                "field_name": field_name,
                "field_description": field_description,
                "column_index": column_index,
            })

    return rows


def _build_summary(records: list[dict], schema_rows: list[dict], return_type: str) -> dict:
    # Column labels derived from records
    col_labels = [_get_period_label(r, return_type) for r in records]

    # Precompute normalized key maps per record for robust lookup
    normalized_records = []
    for r in records:
        norm_map = { _normalize_key(k): k for k in r.keys() }
        normalized_records.append((r, norm_map))

    def lookup_value(rec: dict, norm_map: dict, field_name: str):
        # Direct match first
        if field_name in rec:
            return rec.get(field_name)
        # Normalized match
        nk = _normalize_key(field_name)
        actual_key = norm_map.get(nk)
        if actual_key is not None:
            return rec.get(actual_key)
        # Try simple underscore/space variations
        variants = {
            field_name.replace(" ", "_"),
            field_name.replace("_", " "),
            field_name.lower(),
            field_name.upper(),
        }
        for v in variants:
            if v in rec:
                return rec.get(v)
        return None

    # Group by heading -> category -> optional subcategory
    sections = OrderedDict()
    for row in schema_rows:
        heading = row.get("heading") or ""
        # Normalize category names to match preferences schema (fallback to heading when empty)
        category = _normalize_category(row.get("category"), heading)
        subcategory = row.get("subcategory") or None
        field_name = row.get("field_name") or ""
        field_description = row.get("field_description") or field_name
        column_index = row.get("column_index")

        if heading not in sections:
            sections[heading] = {"heading": heading, "categories": OrderedDict()}
        cats = sections[heading]["categories"]
        if category not in cats:
            cats[category] = {"name": category, "rows": [], "subcategories": OrderedDict()}
        cat_obj = cats[category]

        # Collect values for this field across all records
        values = []
        for rec, nm in normalized_records:
            values.append(lookup_value(rec, nm, field_name))

        row_obj = {
            "description": field_description,
            "key": field_name,
            "column_index": column_index,
            "values": values,
        }

        if return_type == "GSTR-3B" and subcategory:
            subcats = cat_obj["subcategories"]
            if subcategory not in subcats:
                subcats[subcategory] = {"name": subcategory, "rows": []}
            subcats[subcategory]["rows"].append(row_obj)
        else:
            cat_obj["rows"].append(row_obj)

    # Sort rows within each category/subcategory by column_index if available
    for heading_obj in sections.values():
        for cat_key, cat_obj in heading_obj["categories"].items():
            if cat_obj["rows"]:
                cat_obj["rows"].sort(key=lambda r: (r["column_index"] is None, r["column_index"] if r["column_index"] is not None else 0))
            subcats = cat_obj.get("subcategories") or {}
            for sub_key in list(subcats.keys()):
                sub = subcats[sub_key]
                sub["rows"].sort(key=lambda r: (r["column_index"] is None, r["column_index"] if r["column_index"] is not None else 0))

    # Convert OrderedDict to lists for JSON
    sections_list = []
    for heading_obj in sections.values():
        cats_list = []
        for cat_obj in heading_obj["categories"].values():
            item = {"name": cat_obj["name"]}
            # Attach rows or subcategories
            if return_type == "GSTR-3B" and cat_obj.get("subcategories"):
                subs_list = []
                for sub in cat_obj["subcategories"].values():
                    subs_list.append({"name": sub["name"], "rows": sub["rows"]})
                item["subcategories"] = subs_list
            if cat_obj["rows"]:
                item["rows"] = cat_obj["rows"]
            cats_list.append(item)
        sections_list.append({"heading": heading_obj["heading"], "categories": cats_list})

    summary = {"returnType": return_type, "columns": col_labels, "sections": sections_list}
    return _sort_summary_columns(summary)


@app.post("/api/summarize")
async def summarize(payload: dict):
    try:
        records = payload.get("records") or []
        return_type = payload.get("returnType") or "GSTR-3B"
        selected_categories = payload.get("selectedCategories") or None
        books_values = payload.get("booksValues") or {}
        if not isinstance(records, list):
            return JSONResponse(status_code=400, content={"error": "Invalid records payload"})

        schema_rows = _load_schema(return_type)
        # Filter by selected categories if provided
        if selected_categories and isinstance(selected_categories, list):
            selected_set = {str(c).strip() for c in selected_categories}
            def row_cat(row: dict):
                heading = (row.get("heading") or "").strip()
                cat = (row.get("category") or "").strip()
                return _normalize_category(cat, heading)
            schema_rows = [row for row in schema_rows if row_cat(row) in selected_set]
        summary = _build_summary(records, schema_rows, return_type)
        return JSONResponse(content=to_jsonable(summary))
    except Exception as e:
        logger.exception("Failed to summarize records")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/api/schema_categories")
async def schema_categories():
    try:
        result = {}
        for rt in ("GSTR-1", "GSTR-3B"):
            rows = _load_schema(rt)
            cats_set = set()
            for row in rows:
                heading = (row.get("heading") or "").strip()
                cat = _normalize_category(row.get("category"), heading)
                if cat:
                    cats_set.add(cat)
            cats = sorted(cats_set)
            result[rt] = {"categories": cats}
        return JSONResponse(content=to_jsonable(result))
    except Exception as e:
        logger.exception("Failed to load schema categories")
        return JSONResponse(status_code=500, content={"error": str(e)})
async def _extract_any(file: UploadFile, pdf_fn, json_fn):
    """Read an uploaded GSTR-1/3B file and extract it, branching on .pdf vs .json."""
    content = await file.read()
    name = (file.filename or "").lower()
    if name.endswith(".json"):
        return json_fn(json.loads(content))
    with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp:
        tmp.write(content)
        tmp_path = tmp.name
    try:
        return pdf_fn(tmp_path)
    finally:
        os.unlink(tmp_path)    
@app.post("/api/reconcile/gstr1-3b")
async def reconcile_gstr1_gstr3b(
    gstr1: UploadFile = File(...),
    gstr3b: UploadFile = File(...),
    period: str = Form(None),
):
    g1_data  = await _extract_any(gstr1,  extract_gstr1_data,    extract_gstr1_from_json)
    g3b_data = await _extract_any(gstr3b, parse_gstr3b,          extract_gstr3b_from_json)
    result   = reconcile_gstr1_vs_gstr3b(g1_data, g3b_data, period)
    return result

@app.post("/api/reconcile/gstr1-3b/from-json")
async def reconcile_gstr1_gstr3b_from_json(payload: dict):
    """
    Accepts already-extracted GSTR-1 and GSTR-3B dicts (from frontend).
    Body: { "gstr1": {...}, "gstr3b": {...}, "period": "..." }
    """
    g1_data  = payload.get("gstr1", {})
    g3b_data = payload.get("gstr3b", {})
    period   = payload.get("period", "")

    # GSTIN validation
    g1_gstin  = str(g1_data.get("GSTIN") or g1_data.get("gstin") or "").strip()
    g3b_gstin = str(g3b_data.get("gstin") or g3b_data.get("GSTIN") or "").strip()
    if g1_gstin and g3b_gstin and g1_gstin != g3b_gstin:
        return JSONResponse(
            status_code=400,
            content={"error": f"GSTIN mismatch: GSTR-1 has {g1_gstin}, GSTR-3B has {g3b_gstin}"}
        )

    result = reconcile_gstr1_vs_gstr3b(g1_data, g3b_data, period)
    return result


@app.post("/api/reconcile/gstr1-3b/from-pdf")
async def reconcile_from_pdf(
    gstr1_files:  List[UploadFile] = File(...),
    gstr3b_files: List[UploadFile] = File(...),
):
    """
    Accept 1–12 GSTR-1 PDFs/JSONs and 1–12 GSTR-3B PDFs/JSONs.
    Auto-detects periods from extracted data and reconciles per period.
    """
    # Extract all GSTR-1 files
    g1_records = []
    for f in gstr1_files:
        data = await _extract_any(f, extract_gstr1_data, extract_gstr1_from_json)
        data['filename'] = f.filename
        g1_records.append(data)

    # Extract all GSTR-3B files
    g3b_records = []
    for f in gstr3b_files:
        data = await _extract_any(f, parse_gstr3b, extract_gstr3b_from_json)
        data['filename'] = f.filename
        g3b_records.append(data)

    # GSTIN validation — all files must have the same GSTIN
    all_gstins = set()
    for r in g1_records + g3b_records:
        g = str(r.get('GSTIN') or r.get('gstin') or '').strip()
        if g:
            all_gstins.add(g)
    if len(all_gstins) > 1:
        return JSONResponse(status_code=400, content={
            "error": f"GSTIN mismatch across files: {', '.join(all_gstins)}"
        })

    # Auto-detect period from each record
    def get_period(r: dict) -> str:
        tp = r.get('TaxPeriod') or r.get('tax_period') or r.get('period') or ''
        fy = r.get('FinancialYear') or r.get('year') or ''
        if tp and fy:
            return f"{tp} {fy}"
        return tp or r.get('filename', 'Unknown')

    # Build period-keyed dicts
    g1_by_period  = {get_period(r): r for r in g1_records}
    g3b_by_period = {get_period(r): r for r in g3b_records}

    all_periods = sorted(
        set(list(g1_by_period.keys()) + list(g3b_by_period.keys())),
        key=lambda p: (_month_order_index(p.split()[0]) or 99) if p else 99
    )

    period_results = []
    for period in all_periods:
        g1  = g1_by_period.get(period)
        g3b = g3b_by_period.get(period)
        if not g1 or not g3b:
            period_results.append({
                "period": period,
                "gstin": next(iter(all_gstins), ''),
                "overall_status": "Mismatch",
                "total_variance": 0,
                "rows": [{
                    "section": "—",
                    "description": "GSTR-1 missing" if not g1 else "GSTR-3B missing",
                    "gstr1": 0, "gstr3b": 0, "difference": 0, "status": "Missing"
                }]
            })
            continue
        result = reconcile_gstr1_vs_gstr3b(g1, g3b, period)
        period_results.append(result)

    return JSONResponse(content={
        "periods": period_results,
        "gstin": next(iter(all_gstins), ''),
        "months_detected": len(all_periods),
        "g1_extracted":  to_jsonable(g1_records),
        "g3b_extracted": to_jsonable(g3b_records),
    })

@app.post("/api/reconcile/gstr1-3b/excel")
async def reconcile_gstr1_gstr3b_excel(payload: dict):
    """
    Generate and stream the reconciliation Excel report.
    Body: { "periods": [...recon results...], "report_type": "monthly"|"annual" }
    """
    periods = payload.get("periods", [])
    if not periods:
        return JSONResponse(status_code=400, content={"error": "No period data provided."})
    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name
        generate_excel_report(periods, tmp_path)
        report_type = payload.get("report_type", "monthly")
        filename = f"GSTR1_vs_3B_{report_type}_{datetime.datetime.now().strftime('%Y%m%d')}.xlsx"
        with open(tmp_path, "rb") as f:
            content = f.read()
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        logger.error("Excel generation failed: %s", e, exc_info=True)
        return JSONResponse(status_code=500, content={"error": str(e)})
    finally:
        if tmp_path:
            try: os.unlink(tmp_path)
            except: pass
@app.get("/api/reconcile/gstr3b-books/template")
async def download_books_template():
    """Download the Books data Excel template for GSTR-3B vs Books reconciliation."""
    try:
        from backend.core.books_extractor import generate_books_template
        content = generate_books_template()
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=BOOKS_DATA_Template.xlsx"}
        )
    except Exception as e:
        logger.exception("Failed to generate Books template")
        return JSONResponse(status_code=500, content={"error": str(e)})       

@app.get("/api/reconcile/gstr3b-books/sample")
async def download_books_sample():
    """Download a filled-in sample Books workbook (same layout as the template)."""
    try:
        from backend.core.books_extractor import generate_books_sample
        content = generate_books_sample()
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=BOOKS_DATA_Sample.xlsx"}
        )
    except Exception as e:
        logger.exception("Failed to generate Books sample")
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.post("/api/reconcile/gstr3b-books")
async def reconcile_gstr3b_books(
    gstr3b_files: List[UploadFile] = File(...),
    books_file:   UploadFile       = File(...),
):
    """
    Accept 1–12 GSTR-3B PDFs + one Books Excel/CSV.
    Returns reconciliation result dict.
    """
    # Extract all GSTR-3B PDFs
    g3b_records = []
    for f in gstr3b_files:
        data = await _extract_any(f, parse_gstr3b, extract_gstr3b_from_json)
        data['filename'] = f.filename
        g3b_records.append(data)

    # Extract Books file
    books_bytes = await books_file.read()
    try:
        books_data = extract_books(books_bytes, books_file.filename)
    except ValueError as e:
        return JSONResponse(status_code=400, content={"error": str(e)})

    # Optional GSTIN cross-check
    gstin_warning = None
    g3b_gstin   = str(g3b_records[0].get('gstin') or g3b_records[0].get('GSTIN') or '').strip()
    books_gstin = str(books_data.get('gstin') or '').strip()
    if g3b_gstin and books_gstin and g3b_gstin != books_gstin:
        gstin_warning = f"GSTIN mismatch: GSTR-3B={g3b_gstin}, Books={books_gstin}"

    result = reconcile_gstr3b_vs_books(g3b_records, books_data, gstin_warning)
    result['gstin_3b']    = g3b_gstin
    result['gstin_books'] = books_gstin
    result['months_detected'] = len(result.get('monthly', []))
    result['g3b_extracted'] = to_jsonable(g3b_records) 
    return JSONResponse(content=to_jsonable(result))


@app.post("/api/reconcile/gstr3b-books/excel")
async def reconcile_gstr3b_books_excel(payload: dict):
    """
    Generate Excel report from already-reconciled result.
    Body: { result: {...}, gstin_3b: str, gstin_books: str }
    """
    try:
        result      = payload.get('result', {})
        gstin_3b    = payload.get('gstin_3b', '')
        gstin_books = payload.get('gstin_books', '')

        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            tmp_path = tmp.name

        generate_gstr3b_books_excel(result, tmp_path, gstin_3b, gstin_books)

        with open(tmp_path, 'rb') as f:
            content = f.read()
        os.unlink(tmp_path)

        return Response(
            content=content,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename=gstr3b_vs_books.xlsx'}
        )
    except Exception as e:
        logger.exception("Failed to generate GSTR-3B vs Books Excel")
        return JSONResponse(status_code=500, content={"error": str(e)})


# ─── GSTR-2B vs Books ────────────────────────────────────────────────────────
# Reuses the same Books template/extractor as GSTR-3B vs Books (extract_books),
# and the same upload/extract/reconcile/report API shape — see that section
# above for the pattern this mirrors.

@app.get("/api/reconcile/gstr2b-books/template")
async def download_gstr2b_books_template():
    """Download the Books data Excel template — identical format used by
    GSTR-3B vs Books, since both reconciliations consume the same Books upload."""
    try:
        from backend.core.books_extractor import generate_books_template
        content = generate_books_template()
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=BOOKS_DATA_Template.xlsx"}
        )
    except Exception as e:
        logger.exception("Failed to generate Books template")
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.post("/api/reconcile/gstr2b-books")
async def reconcile_gstr2b_books(
    gstr2b_files: List[UploadFile] = File(...),
    books_file:   UploadFile       = File(...),
):
    """
    Accept 1–12 GSTR-2B Excel files (one file = one month, per portal convention)
    + one Books Excel/CSV. Returns reconciliation result dict.
    """
    gstr2b_extracted = []
    for f in gstr2b_files:
        content = await f.read()
        try:
            ext = extract_gstr2b(content, f.filename or "")
        except ValueError as e:
            return JSONResponse(status_code=400, content={"error": f"{f.filename}: {str(e)}"})
        gstr2b_extracted.append(ext)

    gstr2b_data = merge_gstr2b_files(gstr2b_extracted)

    books_bytes = await books_file.read()
    try:
        books_data = extract_books(books_bytes, books_file.filename)
    except ValueError as e:
        return JSONResponse(status_code=400, content={"error": str(e)})

    # GSTIN cross-check (best-effort — Books template has no GSTIN column today,
    # so this only fires when one is actually present in extracted data).
    gstin_warning = None
    gstr2b_gstin = str(gstr2b_data.get("gstin") or "").strip()
    books_gstin = str(books_data.get("gstin") or "").strip()
    if gstr2b_gstin and books_gstin and gstr2b_gstin != books_gstin:
        gstin_warning = f"GSTIN mismatch: GSTR-2B={gstr2b_gstin}, Books={books_gstin}"

    result = reconcile_gstr2b_vs_books(gstr2b_data, books_data, gstin_warning)
    result["gstin_2b"] = gstr2b_gstin
    result["gstin_books"] = books_gstin
    result["months_detected"] = len(result.get("monthly", []))
    result["gstr2b_data"] = gstr2b_data
    result["books_data"] = books_data
    return JSONResponse(content=to_jsonable(result))


@app.post("/api/reconcile/gstr2b-books/excel")
async def reconcile_gstr2b_books_excel(payload: dict):
    """
    Generate Excel report from an already-reconciled result.
    Body: { result: {...}, gstr2b_data: {...}, books_data: {...}, gstin_2b: str, gstin_books: str }
    """
    try:
        result = payload.get("result", {})
        gstr2b_data = payload.get("gstr2b_data", {})
        books_data = payload.get("books_data", {})
        gstin_2b = payload.get("gstin_2b", "")
        gstin_books = payload.get("gstin_books", "")

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp_path = tmp.name

        generate_gstr2b_books_excel(result, tmp_path, gstr2b_data, books_data, gstin_2b, gstin_books)

        with open(tmp_path, "rb") as f:
            content = f.read()
        os.unlink(tmp_path)

        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=gstr2b_vs_books.xlsx"}
        )
    except Exception as e:
        logger.exception("Failed to generate GSTR-2B vs Books Excel")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/api/schema_structure")
async def schema_structure():
    """Return schema organized by heading -> categories -> subcategories preserving order."""
    try:
        def build(rt: str):
            rows = _load_schema(rt)
            sections: "OrderedDict[str, OrderedDict[str, OrderedDict[str, bool]]]" = OrderedDict()
            for r in rows:
                heading = (r.get("heading") or "").strip()
                category = _normalize_category(r.get("category"), heading)
                subcat = r.get("subcategory")
                if isinstance(subcat, str):
                    subcat = _normalize_optional(subcat)
                if heading not in sections:
                    sections[heading] = OrderedDict()
                cats = sections[heading]
                if category not in cats:
                    cats[category] = OrderedDict()
                if subcat:
                    if subcat not in cats[category]:
                        cats[category][subcat] = True

            # Convert to list form preserving order
            out = []
            for heading_name, cats in sections.items():
                cat_list = []
                for cname, subdict in cats.items():
                    subs = list(subdict.keys())
                    item = {"name": cname}
                    if subs:
                        item["subcategories"] = subs
                    cat_list.append(item)
                out.append({"heading": heading_name, "categories": cat_list})
            return out

        return JSONResponse(content={
            "GSTR-1": build("GSTR-1"),
            "GSTR-3B": build("GSTR-3B"),
        })
    except Exception as e:
        logger.exception("Failed to build schema structure")
        return JSONResponse(status_code=500, content={"error": str(e)})


def _safe_sheet_name(name: str, used: set[str]) -> str:
    """Make a string safe and unique as an Excel worksheet name (max 31 chars, no : \\ / ? * [ ]).
    Truncates at the last word boundary where possible, rather than cutting mid-word."""
    cleaned = re.sub(r'[:\\/?*\[\]]', '-', (name or "Sheet").strip()) or "Sheet"
    if len(cleaned) <= 31:
        base = cleaned
    else:
        truncated = cleaned[:31]
        last_space = truncated.rfind(" ")
        base = truncated[:last_space] if last_space >= 10 else truncated
    candidate = base
    n = 2
    while candidate.lower() in used:
        suffix = f" ({n})"
        candidate = base[: 31 - len(suffix)] + suffix
        n += 1
    used.add(candidate.lower())
    return candidate


def _write_table_block(ws, headers: list[str], cols: list[str], rows: list[dict], to_number):
    """Write one table's data (header row, one row per period, bold total row) into a fresh sheet."""
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font
    from openpyxl.worksheet.table import Table, TableStyleInfo

    ws.append(headers)
    header_row_idx = ws.max_row
    for c_idx in range(1, len(headers) + 1):
        ws.cell(row=header_row_idx, column=c_idx).font = Font(bold=True)

    data_start_row = header_row_idx + 1
    first_metric_col_idx = 2
    last_metric_col_idx = len(headers)

    for ci in range(len(cols)):
        row_vals = [cols[ci]]
        for r in rows:
            values = r.get("values", [])
            val = values[ci] if ci < len(values) else None
            row_vals.append(to_number(val))
        ws.append(row_vals)
    data_end_row = ws.max_row

    if data_end_row >= data_start_row and last_metric_col_idx >= first_metric_col_idx:
        totals_row_idx = data_end_row + 1
        total_row_vals = ["Total"]
        for col_idx in range(first_metric_col_idx, last_metric_col_idx + 1):
            col_letter = get_column_letter(col_idx)
            total_row_vals.append(f"=SUM({col_letter}{data_start_row}:{col_letter}{data_end_row})")
        ws.append(total_row_vals)
        for c_idx in range(1, last_metric_col_idx + 1):
            ws.cell(row=totals_row_idx, column=c_idx).font = Font(bold=True)

        for r_idx in range(data_start_row, totals_row_idx + 1):
            for col_idx in range(first_metric_col_idx, last_metric_col_idx + 1):
                ws.cell(row=r_idx, column=col_idx).number_format = "#,##0"

        end_letter = get_column_letter(last_metric_col_idx)
        table_ref = f"A{header_row_idx}:{end_letter}{totals_row_idx}"
        tab = Table(displayName=f"T{abs(hash((ws.title, table_ref))) % 100000}", ref=table_ref)
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False,
        )
        ws.add_table(tab)

    # Reasonable default column widths so data isn't squashed
    ws.column_dimensions["A"].width = 16
    for c_idx in range(2, last_metric_col_idx + 1):
        ws.column_dimensions[get_column_letter(c_idx)].width = 18


def _write_all_table_sheets(wb, summary: dict, used_names: set[str], to_number, prefix: str = ""):
    """Walk a summary's sections/categories/subcategories and write each as its own named sheet."""
    cols = summary.get("columns", [])
    for section in summary.get("sections", []):
        sec_heading = section.get("heading") or ""
        for cat in section.get("categories", []):
            cat_name = cat.get("name") or ""
            if cat.get("subcategories"):
                for sub in cat["subcategories"]:
                    sub_name = (sub.get('name') or '').strip()
                    title = (f"{prefix}{cat_name} - {sub_name}" if sub_name else f"{prefix}{cat_name}").strip(" -") or sec_heading
                    rows = sub.get("rows", [])
                    headers = ["Period"] + [(row.get("description") or row.get("key") or "").strip() for row in rows]
                    ws = wb.create_sheet(title=_safe_sheet_name(title, used_names))
                    _write_table_block(ws, headers, cols, rows, to_number)
            else:
                title = f"{prefix}{cat_name}".strip(" -") or sec_heading
                rows = cat.get("rows", [])
                headers = ["Period"] + [(row.get("description") or row.get("key") or "").strip() for row in rows]
                ws = wb.create_sheet(title=_safe_sheet_name(title, used_names))
                _write_table_block(ws, headers, cols, rows, to_number)


def _apply_category_filter(summary: dict, selected_categories):
    if not (selected_categories and isinstance(selected_categories, list)):
        return summary
    selected_set = {str(c).strip() for c in selected_categories}

    def cat_in_selected(sec_heading: str, cat_name: str):
        return _normalize_category(cat_name or "", sec_heading or "") in selected_set

    filtered_sections = []
    for section in summary.get("sections", []):
        sec_heading = section.get("heading") or ""
        cats = [c for c in section.get("categories", []) if cat_in_selected(sec_heading, c.get("name") or "")]
        if cats:
            new_section = dict(section)
            new_section["categories"] = cats
            filtered_sections.append(new_section)
    summary["sections"] = filtered_sections
    return summary


def _resolve_dataset(dataset: dict) -> dict:
    """Build a (possibly filtered/overridden) summary dict for one records+returnType dataset."""
    records = dataset.get("records") or []
    return_type = dataset.get("returnType") or "GSTR-3B"
    selected_categories = dataset.get("selectedCategories") or None
    if not isinstance(records, list):
        raise ValueError("Invalid records payload")

    schema_rows = _load_schema(return_type)
    if selected_categories and isinstance(selected_categories, list):
        selected_set = {str(c).strip() for c in selected_categories}

        def row_cat(row: dict):
            heading = (row.get("heading") or "").strip()
            cat = (row.get("category") or "").strip()
            return _normalize_category(cat, heading)
        schema_rows = [row for row in schema_rows if row_cat(row) in selected_set]

    summary_override = dataset.get("summaryOverride")
    if isinstance(summary_override, dict) and summary_override.get("sections") and summary_override.get("columns"):
        summary = summary_override
    else:
        summary = _build_summary(records, schema_rows, return_type)

    return _apply_category_filter(summary, selected_categories)


@app.post("/api/tables_excel")
async def tables_excel(payload: dict):
    """Generate an Excel workbook with one named sheet per extracted table.

    Accepts either:
      - a single dataset:  { records, returnType, selectedCategories?, summaryOverride? }
      - multiple datasets: { datasets: [ {records, returnType, ...}, {records, returnType, ...} ] }
        (used by the Reconciliation page to export GSTR-1 + GSTR-3B tables in one workbook)
    """
    try:
        def _to_number(val):
            try:
                if val is None:
                    return None
                if isinstance(val, (int, float)):
                    return float(val)
                if isinstance(val, decimal.Decimal):
                    return float(val)
                if isinstance(val, str):
                    s = val.strip()
                    if s == "":
                        return None
                    return float(s.replace(",", ""))
                return val
            except Exception:
                return val

        datasets = payload.get("datasets")
        if not isinstance(datasets, list) or not datasets:
            datasets = [payload]

        wb = Workbook()
        used_names: set[str] = set()
        for ds in datasets:
            try:
                summary = _resolve_dataset(ds)
            except ValueError as e:
                return JSONResponse(status_code=400, content={"error": str(e)})
            _write_all_table_sheets(wb, summary, used_names, _to_number)

        # The Workbook() constructor creates one default empty sheet — drop it now that
        # real named sheets exist (keep it only if nothing was written, to avoid a corrupt file).
        if len(wb.sheetnames) > 1 and "Sheet" in wb.sheetnames and wb["Sheet"].max_row == 1 and wb["Sheet"].max_column == 1:
            wb.remove(wb["Sheet"])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        headers = {
            "Content-Disposition": "attachment; filename=multi_tables.xlsx",
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        }
        return StreamingResponse(output, headers=headers)
    except Exception as e:
        logger.exception("Failed to create Tables Excel")
        return JSONResponse(status_code=500, content={"error": str(e)})

# Serve frontend (Vite build) as static SPA (mounted last to avoid shadowing /api routes)
try:
    if os.path.exists(DIST_DIR):
        # html=True enables SPA fallback to index.html for unknown paths
        app.mount("/", StaticFiles(directory=DIST_DIR, html=True), name="static")
        logger.info(f"Mounted frontend dist at: {DIST_DIR}")
    else:
        logger.warning(f"Frontend dist not found at {DIST_DIR}; static SPA serving disabled.")
except Exception as e:
    logger.exception(f"Failed to mount StaticFiles for frontend: {e}")