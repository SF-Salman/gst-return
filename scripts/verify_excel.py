import sys
from openpyxl import load_workbook

def main(path: str):
    wb = load_workbook(path, data_only=False)
    ws = wb.active
    # Print header row values
    headers = [c.value for c in ws[1]]
    print("HEADERS:", headers)
    # Print first data row values and formulas in last 3 columns
    if ws.max_row >= 2:
        r = 2
        row_vals = [ws.cell(row=r, column=i).value for i in range(1, ws.max_column+1)]
        print("ROW2 VALUES:", row_vals)
        # Show formulas if present
        for i in range(ws.max_column-2, ws.max_column+1):
            cell = ws.cell(row=r, column=i)
            print(f"R2C{i} -> value={cell.value} number_format={cell.number_format}")

if __name__ == '__main__':
    path = sys.argv[1] if len(sys.argv) > 1 else 'multi_tables.xlsx'
    main(path)